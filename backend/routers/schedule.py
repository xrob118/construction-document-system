"""施工进度横道图 API 路由"""

import os
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from openpyxl import load_workbook

from models import ScheduleTask, Project, GeneratedDoc
from database import get_db, SessionLocal
from services.doc_generator import DocGenerator

router = APIRouter()

# 模板目录
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATES_DIR = os.path.join(BASE_DIR, 'templates')


# ---- Pydantic 请求/响应模型 ----

class ScheduleTaskCreate(BaseModel):
    """创建进度任务请求体"""
    project_id: int
    task_name: str
    parent_id: Optional[int] = None
    start_date: str
    end_date: str
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    progress: int = 0
    responsible: Optional[str] = None
    sort_order: int = 0


class ScheduleTaskUpdate(BaseModel):
    """更新进度任务请求体"""
    task_name: Optional[str] = None
    parent_id: Optional[int] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    actual_start: Optional[str] = None
    actual_end: Optional[str] = None
    progress: Optional[int] = None
    responsible: Optional[str] = None
    sort_order: Optional[int] = None


class ScheduleTaskBatchCreate(BaseModel):
    """批量创建进度任务请求体"""
    project_id: int
    tasks: List[ScheduleTaskCreate]


# ---- CRUD 接口 ----

@router.get("/api/schedule-tasks", summary="获取施工进度任务列表")
def get_schedule_tasks(
    project_id: int = Query(..., description="工程ID"),
    db: Session = Depends(get_db),
):
    """获取指定工程的施工进度任务列表"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    tasks = (
        db.query(ScheduleTask)
        .filter(ScheduleTask.project_id == project_id)
        .order_by(ScheduleTask.sort_order, ScheduleTask.id)
        .all()
    )

    return {
        "project_id": project_id,
        "project_name": project.project_name,
        "items": [
            {
                "id": t.id,
                "project_id": t.project_id,
                "task_name": t.task_name,
                "parent_id": t.parent_id,
                "start_date": t.start_date,
                "end_date": t.end_date,
                "actual_start": t.actual_start,
                "actual_end": t.actual_end,
                "progress": t.progress,
                "responsible": t.responsible,
                "sort_order": t.sort_order,
                "created_at": str(t.created_at) if t.created_at else None,
                "updated_at": str(t.updated_at) if t.updated_at else None,
            }
            for t in tasks
        ],
    }


@router.post("/api/schedule-tasks", summary="创建施工进度任务")
def create_schedule_task(task_data: ScheduleTaskCreate, db: Session = Depends(get_db)):
    """创建单个施工进度任务"""
    project = db.query(Project).filter(Project.id == task_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    task = ScheduleTask(**task_data.model_dump())
    db.add(task)
    db.commit()
    db.refresh(task)
    return {"message": "创建成功", "id": task.id}


@router.post("/api/schedule-tasks/batch", summary="批量创建施工进度任务")
def batch_create_schedule_tasks(batch_data: ScheduleTaskBatchCreate, db: Session = Depends(get_db)):
    """批量创建施工进度任务（先删除该工程已有任务再批量插入）"""
    project = db.query(Project).filter(Project.id == batch_data.project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    # 删除该工程已有任务
    db.query(ScheduleTask).filter(ScheduleTask.project_id == batch_data.project_id).delete()

    # 批量插入
    created_ids = []
    for i, task_data in enumerate(batch_data.tasks):
        task = ScheduleTask(
            project_id=batch_data.project_id,
            task_name=task_data.task_name,
            parent_id=task_data.parent_id,
            start_date=task_data.start_date,
            end_date=task_data.end_date,
            actual_start=task_data.actual_start,
            actual_end=task_data.actual_end,
            progress=task_data.progress,
            responsible=task_data.responsible,
            sort_order=task_data.sort_order if task_data.sort_order else i,
        )
        db.add(task)
        db.flush()
        created_ids.append(task.id)

    db.commit()
    return {"message": "批量创建成功", "count": len(created_ids), "ids": created_ids}


@router.put("/api/schedule-tasks/{task_id}", summary="更新施工进度任务")
def update_schedule_task(task_id: int, task_data: ScheduleTaskUpdate, db: Session = Depends(get_db)):
    """更新施工进度任务"""
    task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    update_data = task_data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(task, key, value)

    db.commit()
    db.refresh(task)
    return {"message": "更新成功", "id": task.id}


@router.delete("/api/schedule-tasks/{task_id}", summary="删除施工进度任务")
def delete_schedule_task(task_id: int, db: Session = Depends(get_db)):
    """删除施工进度任务"""
    task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    db.delete(task)
    db.commit()
    return {"message": "删除成功", "id": task_id}


@router.delete("/api/schedule-tasks/by-project/{project_id}", summary="删除工程下所有进度任务")
def delete_schedule_tasks_by_project(project_id: int, db: Session = Depends(get_db)):
    """删除指定工程下的所有进度任务"""
    count = db.query(ScheduleTask).filter(ScheduleTask.project_id == project_id).delete()
    db.commit()
    return {"message": "删除成功", "count": count}


# ---- Excel 导入功能 ----

def parse_date(value):
    """解析日期值"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    value_str = str(value).strip()
    try:
        dt = datetime.strptime(value_str, '%Y-%m-%d')
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    try:
        dt = datetime.strptime(value_str, '%Y/%m/%d')
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    try:
        dt = datetime.strptime(value_str, '%Y年%m月%d日')
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    # 尝试 Excel 日期序列号（如 46094）
    try:
        num = float(value_str)
        # Excel 日期基准：1900-01-01 对应 1
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=num)
        return dt.strftime('%Y-%m-%d')
    except:
        pass
    return None


def excel_serial_to_date(serial):
    """将 Excel 日期序列号转换为 datetime"""
    try:
        base = datetime(1899, 12, 30)
        return base + timedelta(days=float(serial))
    except:
        return None


def is_cell_filled(cell, target_color='FF00B0F0'):
    """判断单元格是否被目标颜色填充"""
    fill = cell.fill
    if not fill or not fill.fgColor:
        return False
    rgb = fill.fgColor.rgb
    if not rgb or rgb in ('00000000', 'FFFFFFFF', None):
        return False
    # 不指定颜色时，任何非空填充都算
    if target_color is None:
        return fill.patternType == 'solid'
    # 匹配目标颜色（大小写不敏感）
    if isinstance(rgb, str):
        return rgb.upper() == target_color.upper()
    return False


def parse_gantt_template(ws, project):
    """解析横道图模板

    模板格式：
    - 第 1 行：B1~末列为日期序列号
    - A 列：A2~An 为任务名称
    - 横道图条形：A 列右侧单元格被填充色 = 任务工期
    """
    # 1) 读取第 1 行的日期映射：B1=日期1, C1=日期2, ...
    date_map = {}  # col -> datetime
    for col in range(2, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is None:
            continue
        if isinstance(v, datetime):
            date_map[col] = v
        else:
            try:
                num = float(v)
                date_map[col] = excel_serial_to_date(num)
            except:
                # 尝试按字符串解析
                d = parse_date(v)
                if d:
                    date_map[col] = datetime.strptime(d, '%Y-%m-%d')
    if not date_map:
        return []

    # 2) 读取 A 列任务名称
    tasks = []
    for row in range(2, ws.max_row + 1):
        task_name = ws.cell(row=row, column=1).value
        if not task_name or not str(task_name).strip():
            continue
        task_name = str(task_name).strip()

        # 跳过纯说明行（长度>50 通常是说明文字）
        if len(task_name) > 50 and ('\n' in task_name or '：' in task_name):
            continue

        # 3) 查找该行所有被填充的列
        filled_cols = []
        for col in range(2, ws.max_column + 1):
            if is_cell_filled(ws.cell(row=row, column=col)):
                filled_cols.append(col)

        if not filled_cols:
            continue

        # 取最早和最晚的列作为起止日期
        start_col = min(filled_cols)
        end_col = max(filled_cols)

        # 找到对应的日期
        start_date = None
        end_date = None

        # 优先取 col 自身的日期
        if start_col in date_map:
            start_date = date_map[start_col]
        else:
            # 找最近的左侧日期
            for c in sorted(date_map.keys(), reverse=True):
                if c <= start_col:
                    start_date = date_map[c]
                    break

        if end_col in date_map:
            end_date = date_map[end_col]
        else:
            for c in sorted(date_map.keys()):
                if c >= end_col:
                    end_date = date_map[c]
                    break

        if not start_date:
            start_date = datetime.now()
        if not end_date:
            end_date = start_date

        tasks.append({
            'task_name': task_name,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'progress': 0,
            'responsible': None
        })

    return tasks


def reschedule_by_process(tasks_data, project_start_date):
    """按工艺顺序及工艺工期排期（从模板解析的相对偏移平移）

    根据模板中各任务的相对位置关系（谁先谁后、谁和谁并行），
    从工程开始日期出发，按工期依次排期。
    """
    if not tasks_data or not project_start_date:
        return tasks_data

    template_start = None
    for t in tasks_data:
        s = datetime.strptime(t['start_date'], '%Y-%m-%d')
        if not template_start or s < template_start:
            template_start = s

    if not template_start:
        return tasks_data

    project_start = datetime.strptime(project_start_date, '%Y-%m-%d')

    for t in tasks_data:
        orig_start = datetime.strptime(t['start_date'], '%Y-%m-%d')
        orig_end = datetime.strptime(t['end_date'], '%Y-%m-%d')
        start_offset = (orig_start - template_start).days
        end_offset = (orig_end - template_start).days
        t['start_date'] = (project_start + timedelta(days=start_offset)).strftime('%Y-%m-%d')
        t['end_date'] = (project_start + timedelta(days=end_offset)).strftime('%Y-%m-%d')

    return tasks_data


def schedule_from_processes(project, db):
    """从数据库施工工艺数据生成横道图排期

    按模板的汇总工序结构组织，每个汇总工序包含若干工艺库中的细分工艺。
    汇总工序的工期 = 其包含的细分工艺工期之和（串行排列）。
    汇总工序之间的依赖关系按施工逻辑排列。
    不在工艺库中的工序（如设备进场、母线加工及安装）作为补充工序插入。
    """
    from models import Process

    # 读取所有工艺
    processes = db.query(Process).order_by(Process.sort_order, Process.id).all()

    # 工程开始日期
    if project.start_date:
        project_start = datetime.strptime(project.start_date, '%Y-%m-%d')
    else:
        project_start = datetime.now()

    # 定义汇总工序结构
    # 每个汇总工序：名称、包含的工艺名称列表（按顺序串行）、补充工序的工期
    # 补充工序（不在工艺库中的）用 duration 字段指定工期
    summary_structure = [
        {
            'name': '配电房设备基础',
            'process_names': ['基础破碎', '土方开挖', '基础支模', '钢筋绑扎', '接地网焊接',
                              '铁件预埋件焊接', '底板浇筑', '模板安装', '混凝土浇筑',
                              '回填土夯土', '砌体工程', '脚手架搭设', '抹灰工程', '地坪浇筑'],
            'duration': None,  # 由细分工艺工期累加
        },
        {
            'name': '电缆通道',
            'process_names': ['电缆支架安装', '盖板敷设', '土方开挖', '混凝土垫层浇筑',
                              '基础支模', '钢筋绑扎', '接地网焊接', '铁件预埋件焊接',
                              '混凝土浇筑', '土方回填', '盖板安装'],
            'duration': None,  # 由细分工艺工期累加
        },
        {
            'name': '电缆加工',
            'process_names': [],  # 不在工艺库中
            'duration': 23,  # 模板中的工期
            'depends_on_summary': '电缆通道',  # 在电缆通道混凝土浇筑完成后
        },
        {
            'name': '设备进场',
            'process_names': [],  # 不在工艺库中
            'duration': 1,
            'depends_on_summary': '配电房设备基础',  # 在设备基础养护结束后
        },
        {
            'name': '设备安装及调试',
            'process_names': ['变压器安装', '高压柜安装', '低压柜安装', '设备调试'],
            'duration': None,
            'depends_on_summary': '设备进场',  # 在设备进场后
        },
        {
            'name': '母线加工及安装',
            'process_names': ['母线安装'],
            'extra_duration': 11,  # 库里母线安装1天，实际需12天，额外加11天
            'duration': None,
            'depends_on_summary': '设备进场',  # 在设备进场之后开始
        },
        {
            'name': '电缆施放及电缆头制作',
            'process_names': ['电缆敷设', '电缆头制作'],
            'duration': None,
            'depends_on_summary': '电缆加工',  # 在电缆加工完成后
        },
        {
            'name': '试验',
            'process_names': ['电缆试验', '变压器试验', '高压柜试验', '低压柜试验'],
            'duration': None,
            'depends_on_summary': '电缆施放及电缆头制作',
        },
        {
            'name': '营销报验及确定送电时间',
            'process_names': ['送电报验'],
            'duration': None,
            'depends_on_summary': '试验',
        },
        {
            'name': '改造及送电',
            'process_names': [],
            'duration': 1,
            'depends_on_summary': '营销报验及确定送电时间',
        },
    ]

    # 构建工艺名称 -> duration_days 映射
    process_duration_map = {}
    for p in processes:
        process_duration_map[p.name] = p.duration_days or 1

    # 计算每个汇总工序的工期
    for summary in summary_structure:
        if summary['duration'] is not None:
            # 直接指定了工期
            continue

        total_days = 0
        for pname in summary.get('process_names', []):
            total_days += process_duration_map.get(pname, 1)

        # 额外工期（如母线安装库中1天，实际12天）
        if summary.get('extra_duration'):
            total_days += summary['extra_duration']

        # 至少1天
        summary['duration'] = max(1, total_days)

    # 按依赖关系排期
    scheduled = {}  # summary_name -> (start_offset, end_offset)

    def get_summary_schedule(name, visited=None):
        """递归计算汇总工序的排期"""
        if name in scheduled:
            return scheduled[name]

        if visited is None:
            visited = set()
        if name in visited:
            scheduled[name] = (0, 0)
            return scheduled[name]

        visited.add(name)

        # 找到该汇总工序的定义
        summary = None
        for s in summary_structure:
            if s['name'] == name:
                summary = s
                break

        if not summary:
            scheduled[name] = (0, 0)
            return scheduled[name]

        duration = summary['duration'] or 1
        start_offset = 0

        # 依赖的前置工序
        dep_name = summary.get('depends_on_summary')
        if dep_name:
            dep_schedule = get_summary_schedule(dep_name, visited.copy())
            start_offset = dep_schedule[1] + 1  # 前置完成后下一天开始

        end_offset = start_offset + duration - 1
        scheduled[name] = (start_offset, end_offset)
        return scheduled[name]

    # 计算所有汇总工序的排期
    for summary in summary_structure:
        get_summary_schedule(summary['name'])

    # 生成任务列表
    tasks = []
    for i, summary in enumerate(summary_structure):
        if summary['name'] in scheduled:
            start_offset, end_offset = scheduled[summary['name']]
            tasks.append({
                'task_name': summary['name'],
                'start_date': (project_start + timedelta(days=start_offset)).strftime('%Y-%m-%d'),
                'end_date': (project_start + timedelta(days=end_offset)).strftime('%Y-%m-%d'),
                'progress': 0,
                'responsible': None,
                'sort_order': i,
            })

    return tasks


def parse_list_format(ws, project):
    """解析列表格式 Excel

    格式：第 1 行标题 + 数据行
    列：任务名称、开始日期、结束日期、进度、负责人 等
    """
    # 找标题行
    task_name_col = None
    start_date_col = None
    end_date_col = None
    progress_col = None
    responsible_col = None
    start_row = 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        for col_idx, cell_value in enumerate(row, 1):
            if cell_value:
                cell_str = str(cell_value).strip()
                if '任务' in cell_str or '工作' in cell_str:
                    task_name_col = col_idx
                elif '开始' in cell_str:
                    start_date_col = col_idx
                elif '结束' in cell_str:
                    end_date_col = col_idx
                elif '进度' in cell_str or '完成' in cell_str:
                    progress_col = col_idx
                elif '负责人' in cell_str or '责任' in cell_str:
                    responsible_col = col_idx

        if task_name_col:
            start_row = row_idx + 1
            break

    if not task_name_col:
        return []

    tasks = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if not row or len(row) < task_name_col:
            continue
        task_name = row[task_name_col - 1]
        if not task_name or not str(task_name).strip():
            continue

        task_name = str(task_name).strip()
        start_date = row[start_date_col - 1] if start_date_col and len(row) >= start_date_col else None
        end_date = row[end_date_col - 1] if end_date_col and len(row) >= end_date_col else None
        progress = row[progress_col - 1] if progress_col and len(row) >= progress_col else 0
        responsible = row[responsible_col - 1] if responsible_col and len(row) >= responsible_col else None

        start_date_str = parse_date(start_date)
        end_date_str = parse_date(end_date)

        if not start_date_str:
            start_date_str = project.start_date or datetime.now().strftime('%Y-%m-%d')
        if not end_date_str:
            base_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date_str = (base_date + timedelta(days=7)).strftime('%Y-%m-%d')

        progress_val = 0
        if progress:
            try:
                progress_val = min(100, max(0, int(float(progress))))
            except:
                pass

        tasks.append({
            'task_name': task_name,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'progress': progress_val,
            'responsible': str(responsible) if responsible else None
        })

    return tasks


@router.post("/api/schedule-tasks/import-template", summary="从 Excel 模板导入施工进度")
async def import_from_template(project_id: int, use_template: bool = True, db: Session = Depends(get_db)):
    """从 Excel 模板导入施工进度"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    # 查找横道图模板（支持 gantt_chart 前缀或原文件名）
    template_path = None
    if os.path.isdir(TEMPLATES_DIR):
        for f in os.listdir(TEMPLATES_DIR):
            if f.lower().startswith('gantt_chart') or '横道图' in f or '横道图模板' in f:
                template_path = os.path.join(TEMPLATES_DIR, f)
                break

    if not template_path or not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="未找到施工进度横道图模板，请先在模板管理中上传")

    try:
        # 解析 Excel
        wb = load_workbook(template_path, data_only=True)
        ws = wb.active

        # 优先从数据库工艺数据排期
        tasks_data = schedule_from_processes(project, db)

        if not tasks_data:
            # 数据库无工艺数据时，从模板解析
            tasks_data = parse_gantt_template(ws, project)
            if not tasks_data:
                raise HTTPException(status_code=400, detail="未能从模板中解析到有效的任务数据，请检查模板格式")
            # 按模板相对偏移排期
            tasks_data = reschedule_by_process(tasks_data, project.start_date)

        # 删除该工程已有任务
        db.query(ScheduleTask).filter(ScheduleTask.project_id == project_id).delete()

        # 批量插入
        for i, td in enumerate(tasks_data):
            task = ScheduleTask(
                project_id=project_id,
                task_name=td['task_name'],
                parent_id=None,
                start_date=td['start_date'],
                end_date=td['end_date'],
                actual_start=None,
                actual_end=None,
                progress=td['progress'],
                responsible=td['responsible'],
                sort_order=i
            )
            db.add(task)

        db.commit()

        return {
            "message": f"成功导入 {len(tasks_data)} 条任务",
            "count": len(tasks_data),
            "download_url": f"/output/{os.path.basename(template_path)}"
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f'导入模板失败: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f'导入失败: {str(e)}')


@router.post("/api/schedule-tasks/import-excel", summary="上传 Excel 文件导入施工进度")
async def import_from_excel(project_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """上传 Excel 文件导入施工进度

    支持两种格式：
    1. 横道图模板格式：第 1 行是日期序列号，A 列是任务名，单元格颜色表示工期
    2. 列表格式：标题行 + 数据行（任务名/开始日期/结束日期/进度/负责人）
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    # 检查文件类型
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式")

    # 保存临时文件
    import tempfile
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    try:
        content = await file.read()
        temp_file.write(content)
        temp_file.close()

        wb = load_workbook(temp_file.name, data_only=True)
        ws = wb.active

        # 优先从数据库工艺数据排期
        tasks_data = schedule_from_processes(project, db)

        if not tasks_data:
            # 数据库无工艺数据时，从上传的 Excel 解析
            # 先尝试用横道图模板格式解析
            tasks_data = parse_gantt_template(ws, project)

            # 如果横道图格式没解析到数据，尝试列表格式
            if not tasks_data:
                tasks_data = parse_list_format(ws, project)

            if not tasks_data:
                raise HTTPException(status_code=400, detail="未能从文件中解析到有效的任务数据，请检查文件格式")

            # 按模板相对偏移排期
            tasks_data = reschedule_by_process(tasks_data, project.start_date)

        # 删除该工程已有任务
        db.query(ScheduleTask).filter(ScheduleTask.project_id == project_id).delete()

        for i, td in enumerate(tasks_data):
            task = ScheduleTask(
                project_id=project_id,
                task_name=td['task_name'],
                parent_id=td.get('parent_id'),
                start_date=td['start_date'],
                end_date=td['end_date'],
                actual_start=td.get('actual_start'),
                actual_end=td.get('actual_end'),
                progress=td.get('progress', 0),
                responsible=td.get('responsible'),
                sort_order=td.get('sort_order', i)
            )
            db.add(task)

        db.commit()

        return {
            "message": f"成功导入 {len(tasks_data)} 条任务",
            "count": len(tasks_data)
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f'导入 Excel 失败: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f'导入失败: {str(e)}')
    finally:
        try:
            os.unlink(temp_file.name)
        except:
            pass


# ---- 横道图生成（xlsx 文档） ----

class ScheduleGenerateRequest(BaseModel):
    """生成横道图请求体"""
    project_id: int


@router.post("/api/schedule-tasks/generate", summary="生成施工进度横道图文档（xlsx）")
def generate_schedule_chart(request: ScheduleGenerateRequest):
    """从 DB 中读取该工程的进度任务，生成 xlsx 横道图文档并保存到 output/"""
    db = SessionLocal()
    try:
        project = db.query(Project).filter(Project.id == request.project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="工程不存在")

        tasks = (
            db.query(ScheduleTask)
            .filter(ScheduleTask.project_id == request.project_id)
            .order_by(ScheduleTask.sort_order, ScheduleTask.id)
            .all()
        )
        if not tasks:
            raise HTTPException(status_code=400, detail="该工程暂无进度任务，请先从模板导入任务")

        # 构造 generate_gantt_chart 所需的数据格式
        data_dict = {
            "project_name": project.project_name or "未命名工程",
            "tasks": [
                {
                    "name": t.task_name,
                    "start": t.start_date,
                    "end": t.end_date,
                    "progress": t.progress or 0,
                }
                for t in tasks
                if t.start_date and t.end_date
            ],
        }

        if not data_dict["tasks"]:
            raise HTTPException(status_code=400, detail="该工程的进度任务缺少起止日期，无法生成横道图")

        generator = DocGenerator()
        output_path = generator.generate_gantt_chart(data_dict)
        if not output_path:
            raise HTTPException(status_code=500, detail="横道图生成失败")

        filename = os.path.basename(output_path)
        # 保存生成记录
        doc_record = GeneratedDoc(
            project_id=request.project_id,
            doc_type="gantt_chart",
            file_path=output_path,
            pdf_path=None,
        )
        db.add(doc_record)
        db.commit()
        db.refresh(doc_record)
        doc_id = doc_record.id

        return {
            "message": "施工进度横道图生成成功",
            "id": doc_id,
            "filename": filename,
            "file_path": output_path,
            "download_url": f"/api/schedule-tasks/download/{doc_id}",
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f'生成横道图失败: {e}')
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f'生成横道图失败: {str(e)}')
    finally:
        db.close()


@router.get("/api/schedule-tasks/history", summary="获取施工进度横道图生成历史")
def get_schedule_chart_history(
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """获取横道图的生成历史记录"""
    query = db.query(GeneratedDoc).filter(GeneratedDoc.doc_type == "gantt_chart")
    if project_id:
        query = query.filter(GeneratedDoc.project_id == project_id)
    records = query.order_by(GeneratedDoc.created_at.desc()).all()
    return {
        "items": [
            {
                "id": r.id,
                "project_id": r.project_id,
                "doc_type": r.doc_type,
                "file_path": r.file_path,
                "filename": os.path.basename(r.file_path) if r.file_path else None,
                "download_url": f"/api/schedule-tasks/download/{r.id}" if r.file_path else None,
                "created_at": str(r.created_at) if r.created_at else None,
            }
            for r in records
        ]
    }


@router.get("/api/schedule-tasks/download/{doc_id}", summary="下载施工进度横道图文档")
def download_schedule_chart(doc_id: int, db: Session = Depends(get_db)):
    """下载横道图 xlsx 文档"""
    record = db.query(GeneratedDoc).filter(GeneratedDoc.id == doc_id).first()
    if not record or not record.file_path:
        raise HTTPException(status_code=404, detail="文件不存在")

    file_path = record.file_path
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="文件已被删除")

    filename = os.path.basename(file_path)
    return FileResponse(
        path=file_path,
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


class BatchDeleteRequest(BaseModel):
    """批量删除请求体"""
    ids: List[int]


@router.delete("/api/schedule-tasks/history/batch", summary="批量删除施工进度横道图记录")
def batch_delete_schedule_history(request: BatchDeleteRequest, db: Session = Depends(get_db)):
    """批量删除施工进度横道图历史记录"""
    if not request.ids:
        raise HTTPException(status_code=400, detail="ids 不能为空")
    deleted_ids = []
    failed = []
    for doc_id in request.ids:
        try:
            record = db.query(GeneratedDoc).filter(GeneratedDoc.id == doc_id).first()
            if not record:
                failed.append({"id": doc_id, "reason": "记录不存在"})
                continue
            for path_attr in ("file_path", "pdf_path"):
                path_str = getattr(record, path_attr, None)
                if path_str and os.path.exists(path_str):
                    try:
                        os.remove(path_str)
                    except Exception:
                        pass
            db.delete(record)
            deleted_ids.append(doc_id)
        except Exception as e:
            failed.append({"id": doc_id, "reason": str(e)})
    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"提交失败: {str(e)}")

    return {
        "message": "批量删除完成",
        "deleted_ids": deleted_ids,
        "failed": failed,
    }


@router.delete("/api/schedule-tasks/history/{doc_id}", summary="删除施工进度横道图记录")
def delete_schedule_history(doc_id: int, db: Session = Depends(get_db)):
    """删除施工进度横道图历史记录"""
    record = db.query(GeneratedDoc).filter(GeneratedDoc.id == doc_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    for path_attr in ("file_path", "pdf_path"):
        path_str = getattr(record, path_attr, None)
        if path_str and os.path.exists(path_str):
            try:
                os.remove(path_str)
            except Exception:
                pass

    try:
        db.delete(record)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除记录失败: {str(e)}")

    return {"message": "删除成功", "id": doc_id}
