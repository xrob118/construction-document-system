"""工程信息 CRUD API 路由"""

import os
import json
import csv
import shutil
import tempfile
import uuid
from datetime import datetime, timedelta
from urllib.parse import quote
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from io import BytesIO

from models import Project, Process, Worker, ProjectMember, DocTemplate, ProjectApproval, ProjectProcess, ScheduleTask, GeneratedDoc
from database import get_db
from pydantic import BaseModel
import openpyxl

router = APIRouter()

# 模板存储目录
TEMPLATES_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "templates")
os.makedirs(TEMPLATES_DIR, exist_ok=True)

# 工艺导入预览缓存（preview_id -> {items, created_at}）
_IMPORT_CACHE: dict = {}


# ---- Pydantic 请求/响应模型 ----

class ProjectCreate(BaseModel):
    """创建工程请求体"""
    project_code: str
    project_name: str
    subcontractor: Optional[str] = None
    subcontractor_civil: Optional[str] = None
    subcontractor_electric: Optional[str] = None
    project_type: Optional[str] = None
    location: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    description: Optional[str] = None
    # 基本信息
    voltage_level: Optional[str] = None
    line_name: Optional[str] = None
    company_name: Optional[str] = None
    work_task: Optional[str] = None
    # 勘察信息
    survey_unit: Optional[str] = None
    survey_department: Optional[str] = None
    survey_number: Optional[str] = None
    survey_leader: Optional[str] = None
    survey_members: Optional[str] = None
    power_off_range: Optional[str] = None
    live_parts: Optional[str] = None
    danger_points: Optional[str] = None
    # 交底信息
    briefing_host: Optional[str] = None
    # 机具/质控/工期
    equipment_list: Optional[str] = None
    quality_control: Optional[str] = None
    schedule_note: Optional[str] = None
    # 关联
    members: Optional[List[dict]] = None  # [{name, member_type, role}]
    approvals: Optional[List[dict]] = None  # [{role, organization, sort_order}]
    process_ids: Optional[List[int]] = None  # 关联工艺ID列表


class ProjectUpdate(BaseModel):
    """更新工程请求体"""
    project_code: Optional[str] = None
    project_name: Optional[str] = None
    subcontractor: Optional[str] = None
    subcontractor_civil: Optional[str] = None
    subcontractor_electric: Optional[str] = None
    project_type: Optional[str] = None
    location: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    description: Optional[str] = None
    # 基本信息
    voltage_level: Optional[str] = None
    line_name: Optional[str] = None
    company_name: Optional[str] = None
    work_task: Optional[str] = None
    # 勘察信息
    survey_unit: Optional[str] = None
    survey_department: Optional[str] = None
    survey_number: Optional[str] = None
    survey_leader: Optional[str] = None
    survey_members: Optional[str] = None
    power_off_range: Optional[str] = None
    live_parts: Optional[str] = None
    danger_points: Optional[str] = None
    # 交底信息
    briefing_host: Optional[str] = None
    # 机具/质控/工期
    equipment_list: Optional[str] = None
    quality_control: Optional[str] = None
    schedule_note: Optional[str] = None
    # 关联
    members: Optional[List[dict]] = None  # [{name, member_type, role}]
    approvals: Optional[List[dict]] = None  # [{role, organization, sort_order}]
    process_ids: Optional[List[int]] = None  # 关联工艺ID列表


class ProjectResponse(BaseModel):
    """工程响应体"""
    id: int
    project_code: str
    project_name: str
    subcontractor: Optional[str] = None
    subcontractor_civil: Optional[str] = None
    subcontractor_electric: Optional[str] = None
    project_type: Optional[str] = None
    location: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    description: Optional[str] = None
    voltage_level: Optional[str] = None
    line_name: Optional[str] = None
    company_name: Optional[str] = None
    work_task: Optional[str] = None
    survey_unit: Optional[str] = None
    survey_department: Optional[str] = None
    survey_number: Optional[str] = None
    survey_leader: Optional[str] = None
    survey_members: Optional[str] = None
    power_off_range: Optional[str] = None
    live_parts: Optional[str] = None
    danger_points: Optional[str] = None
    briefing_host: Optional[str] = None
    equipment_list: Optional[str] = None
    quality_control: Optional[str] = None
    schedule_note: Optional[str] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None

    class Config:
        from_attributes = True


class ProcessResponse(BaseModel):
    """工艺响应体"""
    id: int
    name: str
    category: Optional[str] = None
    flow_steps: Optional[str] = None
    standards: Optional[str] = None
    description: Optional[str] = None

    class Config:
        from_attributes = True


class WorkerResponse(BaseModel):
    """人员响应体"""
    id: int
    name: str
    role: Optional[str] = None
    role2: Optional[str] = None
    role3: Optional[str] = None
    team: Optional[str] = None
    certification: Optional[str] = None
    certification2: Optional[str] = None
    certification3: Optional[str] = None

    class Config:
        from_attributes = True


# ---- 辅助函数 ----

def _project_to_dict(p, include_detail=False):
    """将 Project 对象转为字典（含人员信息）
    include_detail=True 时返回审批列表、工艺ID列表和工艺详情
    """
    result = {
        "id": p.id,
        "project_code": p.project_code,
        "project_name": p.project_name,
        "subcontractor": p.subcontractor,
        "subcontractor_civil": p.subcontractor_civil,
        "subcontractor_electric": p.subcontractor_electric,
        "project_type": p.project_type,
        "location": p.location,
        "start_date": p.start_date,
        "end_date": p.end_date,
        "description": p.description,
        # 基本信息
        "voltage_level": p.voltage_level,
        "line_name": p.line_name,
        "company_name": p.company_name,
        "work_task": p.work_task,
        # 勘察信息
        "survey_unit": p.survey_unit,
        "survey_department": p.survey_department,
        "survey_number": p.survey_number,
        "survey_leader": p.survey_leader,
        "survey_members": p.survey_members,
        "power_off_range": p.power_off_range,
        "live_parts": p.live_parts,
        "danger_points": p.danger_points,
        # 交底信息
        "briefing_host": p.briefing_host,
        # 机具/质控/工期
        "equipment_list": p.equipment_list,
        "quality_control": p.quality_control,
        "schedule_note": p.schedule_note,
        # 时间戳
        "created_at": str(p.created_at) if p.created_at else None,
        "updated_at": str(p.updated_at) if p.updated_at else None,
        # 人员
        "members": [
            {"id": m.id, "name": m.name, "member_type": m.member_type, "role": m.role}
            for m in p.member_list
        ] if hasattr(p, 'member_list') and p.member_list else [],
    }

    if include_detail:
        # 审批列表
        result["approvals"] = [
            {"id": a.id, "role": a.role, "organization": a.organization, "sort_order": a.sort_order}
            for a in p.approvals
        ] if hasattr(p, 'approvals') and p.approvals else []

        # 工艺ID列表
        result["process_ids"] = [
            pp.process_id for pp in p.process_links
        ] if hasattr(p, 'process_links') and p.process_links else []

        # 工艺详情列表
        process_details = []
        if hasattr(p, 'process_links') and p.process_links:
            for pp in p.process_links:
                proc = pp.process
                if proc:
                    process_details.append({
                        "id": proc.id,
                        "code": proc.code or "",
                        "name": proc.name,
                        "project_type": proc.project_type,
                        "category": proc.category,
                        "sub_category": proc.sub_category,
                        "flow_steps": proc.flow_steps,
                        "duration_days": proc.duration_days,
                        "sort_order": proc.sort_order,
                        "depends_on": proc.depends_on,
                        "standards": proc.standards,
                        "equipment": proc.equipment,
                        "hazards": proc.hazards,
                        "safety_measures": proc.safety_measures,
                        "description": proc.description,
                    })
        result["process_details"] = process_details

    return result


def _save_members(project_id, members_data, db):
    """保存项目人员列表（先删后增）"""
    db.query(ProjectMember).filter(ProjectMember.project_id == project_id).delete()
    for m in members_data:
        member = ProjectMember(
            project_id=project_id,
            name=m.get("name", ""),
            member_type=m.get("member_type", "worker"),
            role=m.get("role", ""),
        )
        db.add(member)
    db.commit()


def _save_approvals(project_id, approvals_data, db):
    """保存项目审批列表（先删后增）"""
    db.query(ProjectApproval).filter(ProjectApproval.project_id == project_id).delete()
    for a in approvals_data:
        approval = ProjectApproval(
            project_id=project_id,
            role=a.get("role", ""),
            organization=a.get("organization", ""),
            sort_order=a.get("sort_order", 0),
        )
        db.add(approval)
    db.flush()


def _save_process_links(project_id, process_ids, db):
    """保存项目工艺关联（先删后增）"""
    db.query(ProjectProcess).filter(ProjectProcess.project_id == project_id).delete()
    for pid in process_ids:
        link = ProjectProcess(
            project_id=project_id,
            process_id=pid,
        )
        db.add(link)
    db.flush()


# ---- 工程信息 CRUD 接口 ----

@router.get("/api/projects", summary="获取工程列表")
def get_projects(
    search: Optional[str] = Query(None, description="搜索关键词（工程名称/编号）"),
    page: int = Query(1, ge=1, description="页码"),
    page_size: int = Query(20, ge=1, le=100, description="每页数量"),
    db: Session = Depends(get_db),
):
    """获取工程列表，支持搜索和分页"""
    query = db.query(Project)

    # 搜索过滤
    if search:
        search_pattern = f"%{search}%"
        query = query.filter(
            (Project.project_name.like(search_pattern)) |
            (Project.project_code.like(search_pattern))
        )

    # 计算总数
    total = query.count()

    # 分页
    offset = (page - 1) * page_size
    projects = query.order_by(Project.created_at.desc()).offset(offset).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [_project_to_dict(p) for p in projects],
    }


# 注意：静态路径路由必须放在 {project_id} 动态路径之前，否则 "preview" 会被当作 project_id 匹配
@router.get("/api/projects/preview/{doc_id}", summary="文档预览")
def preview_document(
    doc_id: int,
    type: Optional[str] = Query("pdf", description="预览类型: pdf=LibreOffice/Word COM 转换的 PDF（与 Word 排版一致）, docx=返回 DOCX 文件"),
    db: Session = Depends(get_db),
):
    """根据文档ID预览生成的文档

    - 默认（type=pdf）：使用 LibreOffice 或 docx2pdf 将 DOCX 转为 PDF，与 Word 排版一致
    - type=docx：返回 DOCX 文件
    """
    doc = db.query(GeneratedDoc).filter(GeneratedDoc.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="文档不存在")

    # PDF 预览：优先 LibreOffice，回退 docx2pdf，再回退数据库 pdf_path
    if type == "pdf":
        if not doc.file_path or not os.path.exists(doc.file_path):
            raise HTTPException(status_code=404, detail="DOCX 源文件不存在，无法预览")
        pdf_path = os.path.splitext(doc.file_path)[0] + ".pdf"

        # 如果 PDF 已存在且比 DOCX 新，直接使用
        if os.path.exists(pdf_path) and os.path.getmtime(pdf_path) >= os.path.getmtime(doc.file_path):
            return FileResponse(
                path=pdf_path,
                media_type="application/pdf",
                filename=os.path.basename(pdf_path),
                content_disposition_type="inline",
            )

        # 方式1：LibreOffice 命令行（稳定可靠）
        try:
            import subprocess
            soffice_path = r"C:\Program Files\LibreOffice\program\soffice.exe"
            if not os.path.exists(soffice_path):
                soffice_path = "soffice"
            result = subprocess.run(
                [soffice_path, "--headless", "--convert-to", "pdf",
                 "--outdir", os.path.dirname(os.path.abspath(doc.file_path)),
                 os.path.abspath(doc.file_path)],
                capture_output=True, timeout=120, errors='replace',
            )
            if result.returncode == 0 and os.path.exists(pdf_path):
                return FileResponse(
                    path=pdf_path,
                    media_type="application/pdf",
                    filename=os.path.basename(pdf_path),
                    content_disposition_type="inline",
                )
        except Exception as e:
            print(f"[preview] LibreOffice 转换失败: {e}", flush=True)

        # 方式2：docx2pdf（Word COM）
        try:
            from docx2pdf import convert
            convert(doc.file_path, pdf_path)
            if os.path.exists(pdf_path):
                return FileResponse(
                    path=pdf_path,
                    media_type="application/pdf",
                    filename=os.path.basename(pdf_path),
                    content_disposition_type="inline",
                )
        except Exception as e:
            print(f"[preview] docx2pdf 转换失败: {e}", flush=True)

        # 回退到数据库中已有的 pdf_path
        if doc.pdf_path and os.path.exists(doc.pdf_path):
            return FileResponse(
                path=doc.pdf_path,
                media_type="application/pdf",
                filename=os.path.basename(doc.pdf_path),
                content_disposition_type="inline",
            )
        raise HTTPException(status_code=503, detail="PDF 转换失败，请稍后再试")

    # 默认（type=docx）—— 返回 DOCX
    if doc.file_path and os.path.exists(doc.file_path):
        return FileResponse(
            path=doc.file_path,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            filename=os.path.basename(doc.file_path),
        )

    # Fallback：如果 DOCX 也不存在但 PDF 存在（极端情况）
    if doc.pdf_path and os.path.exists(doc.pdf_path):
        return FileResponse(
            path=doc.pdf_path,
            media_type="application/pdf",
            filename=os.path.basename(doc.pdf_path),
            content_disposition_type="inline",
        )

    raise HTTPException(status_code=404, detail="文档文件不存在")


@router.get("/api/projects/export", summary="导出工程信息到Excel")
def export_projects(db: Session = Depends(get_db)):
    """导出全部工程信息为Excel文件（5个Sheet）"""
    projects = db.query(Project).order_by(Project.created_at.desc()).all()

    wb = openpyxl.Workbook()

    # ---- Sheet1: 基本信息 ----
    ws1 = wb.active
    ws1.title = "基本信息"
    headers1 = [
        "工程编号", "工程名称", "分包单位", "土建分包单位", "电气分包单位", "工程类别", "工程地点",
        "开工日期", "竣工日期", "工程概况",
        "电压等级", "线路名称/设备双重名称", "编制单位", "工作任务",
        "勘察单位", "勘察部门", "勘察编号", "勘察负责人", "勘察人员",
        "停电范围", "保留的带电部位", "作业现场危险点",
        "交底主持人", "工期说明",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    for row_idx, p in enumerate(projects, start=2):
        row_data = [
            p.project_code or "",
            p.project_name or "",
            p.subcontractor or "",
            p.subcontractor_civil or "",
            p.subcontractor_electric or "",
            p.project_type or "",
            p.location or "",
            p.start_date or "",
            p.end_date or "",
            p.description or "",
            p.voltage_level or "",
            p.line_name or "",
            p.company_name or "",
            p.work_task or "",
            p.survey_unit or "",
            p.survey_department or "",
            p.survey_number or "",
            p.survey_leader or "",
            p.survey_members or "",
            p.power_off_range or "",
            p.live_parts or "",
            p.danger_points or "",
            p.briefing_host or "",
            p.schedule_note or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col, value=val)

    # ---- Sheet2: 人员信息 ----
    ws2 = wb.create_sheet("人员信息")
    headers2 = ["工程编号", "姓名", "人员类型", "职务/工种"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    row_idx = 2
    for p in projects:
        if hasattr(p, 'member_list') and p.member_list:
            for m in p.member_list:
                ws2.cell(row=row_idx, column=1, value=p.project_code or "")
                ws2.cell(row=row_idx, column=2, value=m.name or "")
                ws2.cell(row=row_idx, column=3, value=m.member_type or "")
                ws2.cell(row=row_idx, column=4, value=m.role or "")
                row_idx += 1

    # ---- Sheet3: 机具清单 ----
    ws3 = wb.create_sheet("机具清单")
    headers3 = ["工程编号", "名称", "编号", "单位", "数量"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)

    row_idx = 2
    for p in projects:
        if p.equipment_list:
            try:
                eq_list = json.loads(p.equipment_list)
                for eq in eq_list:
                    ws3.cell(row=row_idx, column=1, value=p.project_code or "")
                    ws3.cell(row=row_idx, column=2, value=eq.get("name", ""))
                    ws3.cell(row=row_idx, column=3, value=eq.get("code", ""))
                    ws3.cell(row=row_idx, column=4, value=eq.get("unit", ""))
                    ws3.cell(row=row_idx, column=5, value=eq.get("quantity", ""))
                    row_idx += 1
            except (json.JSONDecodeError, TypeError):
                pass

    # ---- Sheet4: 质量控制点 ----
    ws4 = wb.create_sheet("质量控制点")
    headers4 = ["工程编号", "分项工程", "分部工程", "检验依据", "检验方法", "负责人", "检验记录"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)

    row_idx = 2
    for p in projects:
        if p.quality_control:
            try:
                qc_list = json.loads(p.quality_control)
                for qc in qc_list:
                    ws4.cell(row=row_idx, column=1, value=p.project_code or "")
                    ws4.cell(row=row_idx, column=2, value=qc.get("project", ""))
                    ws4.cell(row=row_idx, column=3, value=qc.get("sub_project", ""))
                    ws4.cell(row=row_idx, column=4, value=qc.get("basis", ""))
                    ws4.cell(row=row_idx, column=5, value=qc.get("method", ""))
                    ws4.cell(row=row_idx, column=6, value=qc.get("responsible", ""))
                    ws4.cell(row=row_idx, column=7, value=qc.get("record", ""))
                    row_idx += 1
            except (json.JSONDecodeError, TypeError):
                pass

    # ---- Sheet5: 审批信息 ----
    ws5 = wb.create_sheet("审批信息")
    headers5 = ["工程编号", "角色", "单位", "排序"]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)

    row_idx = 2
    for p in projects:
        if hasattr(p, 'approvals') and p.approvals:
            for a in p.approvals:
                ws5.cell(row=row_idx, column=1, value=p.project_code or "")
                ws5.cell(row=row_idx, column=2, value=a.role or "")
                ws5.cell(row=row_idx, column=3, value=a.organization or "")
                ws5.cell(row=row_idx, column=4, value=a.sort_order if a.sort_order is not None else 0)
                row_idx += 1

    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('工程信息_导出_' + datetime.now().strftime('%Y%m%d') + '.xlsx')}"},
    )


@router.get("/api/projects/import-template", summary="下载工程导入模板")
def download_import_template():
    """生成空模板Excel（含表头+示例行+校验说明）"""
    wb = openpyxl.Workbook()

    # ---- Sheet1: 基本信息 ----
    ws1 = wb.active
    ws1.title = "基本信息"
    headers1 = [
        "工程编号", "工程名称", "分包单位", "土建分包单位", "电气分包单位", "工程类别", "工程地点",
        "开工日期", "竣工日期", "工程概况",
        "电压等级", "线路名称/设备双重名称", "编制单位", "工作任务",
        "勘察单位", "勘察部门", "勘察编号", "勘察负责人", "勘察人员",
        "停电范围", "保留的带电部位", "作业现场危险点",
        "交底主持人", "工期说明",
    ]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    # 示例行
    example1 = [
        "GC-2026-001", "10kV配电房新建工程", "土建:XX建设有限公司; 电气:YY电气公司", "XX建设有限公司", "YY电气公司", "配电房",
        "XX市XX区XX路", "2026-01-01", "2026-06-30", "新建10kV配电房一座",
        "10kV", "XX线XX杆", "XX电力工程公司", "配电房土建及电气安装",
        "XX勘察设计院", "电力勘察所", "KC-2026-001", "张三", "李四,王五",
        "XX线XX段停电", "XX线带电", "深基坑、高处作业",
        "李工", "总工期180天",
    ]
    for col, val in enumerate(example1, 1):
        ws1.cell(row=2, column=col, value=val)

    # 校验说明
    ws1.cell(row=4, column=1, value="校验说明：")
    ws1.cell(row=5, column=1, value="1. 工程编号（project_code）必填，已存在则覆盖更新")
    ws1.cell(row=6, column=1, value="2. 工程名称（project_name）必填")
    ws1.cell(row=7, column=1, value="3. 日期格式：YYYY-MM-DD（如 2026-01-01）")

    # ---- Sheet2: 人员信息 ----
    ws2 = wb.create_sheet("人员信息")
    headers2 = ["工程编号", "姓名", "人员类型", "职务/工种"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    example2 = ["GC-2026-001", "张建国", "manager", "项目经理"]
    for col, val in enumerate(example2, 1):
        ws2.cell(row=2, column=col, value=val)
    example2b = ["GC-2026-001", "刘施工", "worker", "施工员"]
    for col, val in enumerate(example2b, 1):
        ws2.cell(row=3, column=col, value=val)

    # ---- Sheet3: 机具清单 ----
    ws3 = wb.create_sheet("机具清单")
    headers3 = ["工程编号", "名称", "编号", "单位", "数量"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    example3 = ["GC-2026-001", "挖掘机", "EQ-001", "台", 1]
    for col, val in enumerate(example3, 1):
        ws3.cell(row=2, column=col, value=val)

    # ---- Sheet4: 质量控制点 ----
    ws4 = wb.create_sheet("质量控制点")
    headers4 = ["工程编号", "分项工程", "分部工程", "检验依据", "检验方法", "负责人", "检验记录"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    example4 = ["GC-2026-001", "隐蔽工程", "平面布置", "设计图纸", "跟踪检查", "施工负责人", "评级记录"]
    for col, val in enumerate(example4, 1):
        ws4.cell(row=2, column=col, value=val)

    # ---- Sheet5: 审批信息 ----
    ws5 = wb.create_sheet("审批信息")
    headers5 = ["工程编号", "角色", "单位", "排序"]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    example5 = ["GC-2026-001", "编制", "XX电力工程公司", 1]
    for col, val in enumerate(example5, 1):
        ws5.cell(row=2, column=col, value=val)
    example5b = ["GC-2026-001", "审核", "XX监理公司", 2]
    for col, val in enumerate(example5b, 1):
        ws5.cell(row=3, column=col, value=val)
    example5c = ["GC-2026-001", "批准", "XX供电公司", 3]
    for col, val in enumerate(example5c, 1):
        ws5.cell(row=4, column=col, value=val)

    # 调整列宽
    for ws in [ws1, ws2, ws3, ws4, ws5]:
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('工程导入模板.xlsx')}"},
    )


@router.get("/api/projects/{project_id}", summary="获取单个工程详情")
def get_project(project_id: int, db: Session = Depends(get_db)):
    """获取单个工程的详细信息（含审批、工艺关联）"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")
    return _project_to_dict(project, include_detail=True)


@router.post("/api/projects", summary="创建新工程")
def create_project(project_data: ProjectCreate, db: Session = Depends(get_db)):
    """创建新的工程记录"""
    # 检查工程编号是否重复
    existing = db.query(Project).filter(Project.project_code == project_data.project_code).first()
    if existing:
        raise HTTPException(status_code=400, detail="工程编号已存在")

    members_data = project_data.members or []
    approvals_data = project_data.approvals or []
    process_ids = project_data.process_ids or []
    project_dict = project_data.model_dump(exclude={"members", "approvals", "process_ids"})

    # 兼容逻辑：如果传入了 subcontractor_civil 或 subcontractor_electric，同步更新 subcontractor
    civil = project_data.subcontractor_civil
    electric = project_data.subcontractor_electric
    if civil or electric:
        parts = []
        if civil:
            parts.append(f"土建:{civil}")
        if electric:
            parts.append(f"电气:{electric}")
        project_dict["subcontractor"] = "; ".join(parts)

    project = Project(**project_dict)
    db.add(project)
    db.commit()
    db.refresh(project)

    # 保存人员
    if members_data:
        _save_members(project.id, members_data, db)

    # 保存审批
    if approvals_data:
        _save_approvals(project.id, approvals_data, db)
        db.commit()

    # 保存工艺关联
    if process_ids:
        _save_process_links(project.id, process_ids, db)
        db.commit()

    return _project_to_dict(project, include_detail=True)


@router.put("/api/projects/{project_id}", summary="更新工程")
def update_project(project_id: int, project_data: ProjectUpdate, db: Session = Depends(get_db)):
    """更新工程信息"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    # 如果更新了工程编号，检查是否重复
    if project_data.project_code and project_data.project_code != project.project_code:
        existing = db.query(Project).filter(Project.project_code == project_data.project_code).first()
        if existing:
            raise HTTPException(status_code=400, detail="工程编号已存在")

    # 只更新提供的字段（排除 members, approvals, process_ids）
    update_data = project_data.model_dump(exclude_unset=True, exclude={"members", "approvals", "process_ids"})

    # 兼容逻辑：如果传入了 subcontractor_civil 或 subcontractor_electric，同步更新 subcontractor
    civil = project_data.subcontractor_civil
    electric = project_data.subcontractor_electric
    if civil or electric:
        parts = []
        if civil:
            parts.append(f"土建:{civil}")
        if electric:
            parts.append(f"电气:{electric}")
        update_data["subcontractor"] = "; ".join(parts)

    for key, value in update_data.items():
        setattr(project, key, value)

    db.commit()
    db.refresh(project)

    # 更新人员
    if project_data.members is not None:
        _save_members(project_id, project_data.members, db)

    # 更新审批（先删后增）
    if project_data.approvals is not None:
        _save_approvals(project_id, project_data.approvals, db)
        db.commit()

    # 更新工艺关联（先删后增）
    if project_data.process_ids is not None:
        _save_process_links(project_id, project_data.process_ids, db)
        db.commit()

    return _project_to_dict(project, include_detail=True)


@router.delete("/api/projects/{project_id}", summary="删除工程")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    """删除工程记录"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    db.delete(project)
    db.commit()
    return {"message": "删除成功", "id": project_id}


# ---- Excel 导入/导出接口 ----

def _parse_date_value(value):
    """解析日期值，返回 YYYY-MM-DD 字符串或 None"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    value_str = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日'):
        try:
            dt = datetime.strptime(value_str, fmt)
            return dt.strftime('%Y-%m-%d')
        except (ValueError, TypeError):
            continue
    return None


def _validate_date(value):
    """校验日期格式，返回 (是否合法, 解析后的字符串)"""
    parsed = _parse_date_value(value)
    return parsed is not None, parsed


@router.post("/api/projects/import", summary="Excel导入工程信息")
async def import_projects(
    file: UploadFile = File(..., description="Excel文件(.xlsx/.xls)"),
    db: Session = Depends(get_db),
):
    """上传Excel文件批量导入工程信息

    Sheet1: 基本信息
    Sheet2: 人员信息
    Sheet3: 机具清单
    Sheet4: 质量控制点
    Sheet5: 审批信息

    数据校验：project_code必填且唯一（已存在则覆盖），project_name必填，日期格式校验
    """
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 格式文件")

    content = await file.read()

    try:
        tmp_path = os.path.join(tempfile.gettempdir(), f"upload_{filename}")
        with open(tmp_path, "wb") as f:
            f.write(content)
        wb = openpyxl.load_workbook(tmp_path, read_only=True)

        errors = []
        total = 0
        success = 0

        # ---- Sheet1: 基本信息 ----
        ws1 = wb.worksheets[0] if len(wb.worksheets) > 0 else None
        if not ws1:
            wb.close()
            raise HTTPException(status_code=400, detail="Excel文件至少需要一个Sheet")

        rows1 = list(ws1.iter_rows(values_only=True))
        if len(rows1) < 2:
            wb.close()
            raise HTTPException(status_code=400, detail="Sheet1至少需要包含表头和一行数据")

        header1 = [str(h).strip() if h else "" for h in rows1[0]]
        col_map1 = {}
        for i, h in enumerate(header1):
            if "工程编号" in h or "project_code" in h.lower():
                col_map1["project_code"] = i
            elif "工程名称" in h or "project_name" in h.lower():
                col_map1["project_name"] = i
            elif "分包单位" in h or "subcontractor" in h.lower():
                col_map1["subcontractor"] = i
            elif "土建分包" in h or "subcontractor_civil" in h.lower():
                col_map1["subcontractor_civil"] = i
            elif "电气分包" in h or "subcontractor_electric" in h.lower():
                col_map1["subcontractor_electric"] = i
            elif "工程类别" in h or "工程类型" in h:
                col_map1["project_type"] = i
            elif "工程地点" in h or "地点" in h:
                col_map1["location"] = i
            elif "开工日期" in h or "开始日期" in h:
                col_map1["start_date"] = i
            elif "竣工日期" in h or "结束日期" in h:
                col_map1["end_date"] = i
            elif "工程概况" in h or "概况" in h or "描述" in h:
                col_map1["description"] = i
            elif "电压等级" in h:
                col_map1["voltage_level"] = i
            elif "线路名称" in h or "设备双重名称" in h:
                col_map1["line_name"] = i
            elif "编制单位" in h:
                col_map1["company_name"] = i
            elif "工作任务" in h:
                col_map1["work_task"] = i
            elif "勘察单位" in h:
                col_map1["survey_unit"] = i
            elif "勘察部门" in h:
                col_map1["survey_department"] = i
            elif "勘察编号" in h:
                col_map1["survey_number"] = i
            elif "勘察负责人" in h:
                col_map1["survey_leader"] = i
            elif "勘察人员" in h:
                col_map1["survey_members"] = i
            elif "停电范围" in h:
                col_map1["power_off_range"] = i
            elif "带电部位" in h or "保留带电" in h:
                col_map1["live_parts"] = i
            elif "危险点" in h or "作业现场危险点" in h:
                col_map1["danger_points"] = i
            elif "交底主持人" in h or "主持人" in h:
                col_map1["briefing_host"] = i
            elif "工期说明" in h:
                col_map1["schedule_note"] = i

        if "project_code" not in col_map1:
            wb.close()
            raise HTTPException(status_code=400, detail='Sheet1表头中未找到"工程编号"列')
        if "project_name" not in col_map1:
            wb.close()
            raise HTTPException(status_code=400, detail='Sheet1表头中未找到"工程名称"列')

        # ---- Sheet2: 人员信息 ----
        ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else None
        rows2 = list(ws2.iter_rows(values_only=True)) if ws2 else []
        header2 = [str(h).strip() if h else "" for h in rows2[0]] if len(rows2) > 0 else []
        col_map2 = {}
        for i, h in enumerate(header2):
            if "工程编号" in h or "project_code" in h.lower():
                col_map2["project_code"] = i
            elif "姓名" in h or "名字" in h:
                col_map2["name"] = i
            elif "人员类型" in h or "类型" in h:
                col_map2["member_type"] = i
            elif "职务" in h or "角色" in h or "工种" in h:
                col_map2["role"] = i

        # 按 project_code 分组人员
        members_by_code = {}
        for row_idx, row in enumerate(rows2[1:], start=2) if len(rows2) > 1 else []:
            if not row or all(not cell for cell in row):
                continue
            pc = str(row[col_map2["project_code"]]).strip() if col_map2.get("project_code") is not None and col_map2["project_code"] < len(row) and row[col_map2["project_code"]] else ""
            name = str(row[col_map2["name"]]).strip() if col_map2.get("name") is not None and col_map2["name"] < len(row) and row[col_map2["name"]] else ""
            if not pc or not name:
                continue
            member_type = str(row[col_map2["member_type"]]).strip() if col_map2.get("member_type") is not None and col_map2["member_type"] < len(row) and row[col_map2["member_type"]] else "worker"
            role = str(row[col_map2["role"]]).strip() if col_map2.get("role") is not None and col_map2["role"] < len(row) and row[col_map2["role"]] else ""
            members_by_code.setdefault(pc, []).append({"name": name, "member_type": member_type, "role": role})

        # ---- Sheet3: 机具清单 ----
        ws3 = wb.worksheets[2] if len(wb.worksheets) > 2 else None
        rows3 = list(ws3.iter_rows(values_only=True)) if ws3 else []
        header3 = [str(h).strip() if h else "" for h in rows3[0]] if len(rows3) > 0 else []
        col_map3 = {}
        for i, h in enumerate(header3):
            if "工程编号" in h or "project_code" in h.lower():
                col_map3["project_code"] = i
            elif "名称" in h or "机具名称" in h:
                col_map3["name"] = i
            elif "编号" in h or "机具编号" in h:
                col_map3["code"] = i
            elif "单位" in h:
                col_map3["unit"] = i
            elif "数量" in h:
                col_map3["quantity"] = i

        equipment_by_code = {}
        for row_idx, row in enumerate(rows3[1:], start=2) if len(rows3) > 1 else []:
            if not row or all(not cell for cell in row):
                continue
            pc = str(row[col_map3["project_code"]]).strip() if col_map3.get("project_code") is not None and col_map3["project_code"] < len(row) and row[col_map3["project_code"]] else ""
            if not pc:
                continue
            eq_name = str(row[col_map3["name"]]).strip() if col_map3.get("name") is not None and col_map3["name"] < len(row) and row[col_map3["name"]] else ""
            eq_code = str(row[col_map3["code"]]).strip() if col_map3.get("code") is not None and col_map3["code"] < len(row) and row[col_map3["code"]] else ""
            eq_unit = str(row[col_map3["unit"]]).strip() if col_map3.get("unit") is not None and col_map3["unit"] < len(row) and row[col_map3["unit"]] else ""
            eq_qty = row[col_map3["quantity"]] if col_map3.get("quantity") is not None and col_map3["quantity"] < len(row) and row[col_map3["quantity"]] else ""
            if eq_name:
                equipment_by_code.setdefault(pc, []).append({
                    "name": eq_name, "code": eq_code, "unit": eq_unit,
                    "quantity": eq_qty,
                })

        # ---- Sheet4: 质量控制点 ----
        ws4 = wb.worksheets[3] if len(wb.worksheets) > 3 else None
        rows4 = list(ws4.iter_rows(values_only=True)) if ws4 else []
        header4 = [str(h).strip() if h else "" for h in rows4[0]] if len(rows4) > 0 else []
        col_map4 = {}
        for i, h in enumerate(header4):
            if "工程编号" in h or "project_code" in h.lower():
                col_map4["project_code"] = i
            elif "分项工程" in h or "项目" in h:
                col_map4["project"] = i
            elif "分部工程" in h or "子项目" in h:
                col_map4["sub_project"] = i
            elif "依据" in h or "检验依据" in h:
                col_map4["basis"] = i
            elif "方法" in h or "检验方法" in h:
                col_map4["method"] = i
            elif "负责人" in h or "责任" in h:
                col_map4["responsible"] = i
            elif "记录" in h or "检验记录" in h:
                col_map4["record"] = i

        qc_by_code = {}
        for row_idx, row in enumerate(rows4[1:], start=2) if len(rows4) > 1 else []:
            if not row or all(not cell for cell in row):
                continue
            pc = str(row[col_map4["project_code"]]).strip() if col_map4.get("project_code") is not None and col_map4["project_code"] < len(row) and row[col_map4["project_code"]] else ""
            if not pc:
                continue
            qc_project = str(row[col_map4["project"]]).strip() if col_map4.get("project") is not None and col_map4["project"] < len(row) and row[col_map4["project"]] else ""
            qc_sub = str(row[col_map4["sub_project"]]).strip() if col_map4.get("sub_project") is not None and col_map4["sub_project"] < len(row) and row[col_map4["sub_project"]] else ""
            qc_basis = str(row[col_map4["basis"]]).strip() if col_map4.get("basis") is not None and col_map4["basis"] < len(row) and row[col_map4["basis"]] else ""
            qc_method = str(row[col_map4["method"]]).strip() if col_map4.get("method") is not None and col_map4["method"] < len(row) and row[col_map4["method"]] else ""
            qc_resp = str(row[col_map4["responsible"]]).strip() if col_map4.get("responsible") is not None and col_map4["responsible"] < len(row) and row[col_map4["responsible"]] else ""
            qc_record = str(row[col_map4["record"]]).strip() if col_map4.get("record") is not None and col_map4["record"] < len(row) and row[col_map4["record"]] else ""
            if qc_project:
                qc_by_code.setdefault(pc, []).append({
                    "project": qc_project, "sub_project": qc_sub, "basis": qc_basis,
                    "method": qc_method, "responsible": qc_resp, "record": qc_record,
                })

        # ---- Sheet5: 审批信息 ----
        ws5 = wb.worksheets[4] if len(wb.worksheets) > 4 else None
        rows5 = list(ws5.iter_rows(values_only=True)) if ws5 else []
        header5 = [str(h).strip() if h else "" for h in rows5[0]] if len(rows5) > 0 else []
        col_map5 = {}
        for i, h in enumerate(header5):
            if "工程编号" in h or "project_code" in h.lower():
                col_map5["project_code"] = i
            elif "角色" in h or "审批角色" in h:
                col_map5["role"] = i
            elif "单位" in h or "审批单位" in h:
                col_map5["organization"] = i
            elif "排序" in h or "顺序" in h:
                col_map5["sort_order"] = i

        approvals_by_code = {}
        for row_idx, row in enumerate(rows5[1:], start=2) if len(rows5) > 1 else []:
            if not row or all(not cell for cell in row):
                continue
            pc = str(row[col_map5["project_code"]]).strip() if col_map5.get("project_code") is not None and col_map5["project_code"] < len(row) and row[col_map5["project_code"]] else ""
            if not pc:
                continue
            ap_role = str(row[col_map5["role"]]).strip() if col_map5.get("role") is not None and col_map5["role"] < len(row) and row[col_map5["role"]] else ""
            ap_org = str(row[col_map5["organization"]]).strip() if col_map5.get("organization") is not None and col_map5["organization"] < len(row) and row[col_map5["organization"]] else ""
            ap_sort = row[col_map5["sort_order"]] if col_map5.get("sort_order") is not None and col_map5["sort_order"] < len(row) and row[col_map5["sort_order"]] else 0
            try:
                ap_sort = int(ap_sort)
            except (ValueError, TypeError):
                ap_sort = 0
            if ap_role:
                approvals_by_code.setdefault(pc, []).append({
                    "role": ap_role, "organization": ap_org, "sort_order": ap_sort,
                })

        wb.close()
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        # ---- 逐行处理 Sheet1 数据 ----
        for row_idx, row in enumerate(rows1[1:], start=2):
            if not row or all(not cell for cell in row):
                continue
            total += 1

            def _get_val(key):
                if key not in col_map1:
                    return None
                idx = col_map1[key]
                val = row[idx] if idx < len(row) and row[idx] is not None else None
                return str(val).strip() if val is not None else None

            project_code = _get_val("project_code") or ""
            project_name = _get_val("project_name") or ""

            # 校验
            if not project_code:
                errors.append({"row": row_idx, "field": "project_code", "message": "工程编号不能为空"})
                continue
            if not project_name:
                errors.append({"row": row_idx, "field": "project_name", "message": "工程名称不能为空"})
                continue

            # 日期校验
            start_date_raw = _get_val("start_date")
            end_date_raw = _get_val("end_date")
            start_date = _parse_date_value(start_date_raw) if start_date_raw else None
            end_date = _parse_date_value(end_date_raw) if end_date_raw else None
            if start_date_raw and not start_date:
                errors.append({"row": row_idx, "field": "start_date", "message": f"开工日期格式错误: {start_date_raw}"})
                continue
            if end_date_raw and not end_date:
                errors.append({"row": row_idx, "field": "end_date", "message": f"竣工日期格式错误: {end_date_raw}"})
                continue

            # 构建工程数据
            subcontractor_civil = _get_val("subcontractor_civil")
            subcontractor_electric = _get_val("subcontractor_electric")
            # 兼容逻辑：如果有新列就用新列，否则回退到旧的 subcontractor 列
            if not subcontractor_civil and not subcontractor_electric:
                # 回退：从旧的 subcontractor 列取值
                subcontractor_val = _get_val("subcontractor")
            else:
                # 有新列时，同步更新 subcontractor 为组合格式
                parts = []
                if subcontractor_civil:
                    parts.append(f"土建:{subcontractor_civil}")
                if subcontractor_electric:
                    parts.append(f"电气:{subcontractor_electric}")
                subcontractor_val = "; ".join(parts) if parts else _get_val("subcontractor")

            project_values = {
                "project_code": project_code,
                "project_name": project_name,
                "subcontractor": subcontractor_val,
                "subcontractor_civil": subcontractor_civil,
                "subcontractor_electric": subcontractor_electric,
                "project_type": _get_val("project_type"),
                "location": _get_val("location"),
                "start_date": start_date,
                "end_date": end_date,
                "description": _get_val("description"),
                "voltage_level": _get_val("voltage_level"),
                "line_name": _get_val("line_name"),
                "company_name": _get_val("company_name"),
                "work_task": _get_val("work_task"),
                "survey_unit": _get_val("survey_unit"),
                "survey_department": _get_val("survey_department"),
                "survey_number": _get_val("survey_number"),
                "survey_leader": _get_val("survey_leader"),
                "survey_members": _get_val("survey_members"),
                "power_off_range": _get_val("power_off_range"),
                "live_parts": _get_val("live_parts"),
                "danger_points": _get_val("danger_points"),
                "briefing_host": _get_val("briefing_host"),
                "schedule_note": _get_val("schedule_note"),
            }

            # 机具清单
            eq_list = equipment_by_code.get(project_code, [])
            if eq_list:
                project_values["equipment_list"] = json.dumps(eq_list, ensure_ascii=False)

            # 质量控制点
            qc_list = qc_by_code.get(project_code, [])
            if qc_list:
                project_values["quality_control"] = json.dumps(qc_list, ensure_ascii=False)

            # 已存在则覆盖
            existing = db.query(Project).filter(Project.project_code == project_code).first()
            if existing:
                for key, value in project_values.items():
                    if value is not None:
                        setattr(existing, key, value)
                project_obj = existing
            else:
                project_obj = Project(**project_values)
                db.add(project_obj)

            db.flush()

            # 人员
            members_data = members_by_code.get(project_code, [])
            if members_data:
                _save_members(project_obj.id, members_data, db)

            # 审批
            approvals_data = approvals_by_code.get(project_code, [])
            if approvals_data:
                _save_approvals(project_obj.id, approvals_data, db)

            success += 1

        db.commit()

        return {
            "total": total,
            "success": success,
            "errors": errors,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")

# ---- 施工进度自动排期接口 ----

@router.post("/api/projects/{project_id}/auto-schedule", summary="施工进度自动排期")
def auto_schedule(project_id: int, db: Session = Depends(get_db)):
    """根据工程关联的施工工艺自动生成施工进度排期

    规则：
    a. 从工程 start_date 开始
    b. 无依赖的工艺可以并行，取工期最长的作为后续工艺的开始时间
    c. 有依赖的工艺，在所有前置工艺完成后开始
    d. 土建类工艺全部完成后才能开始电气类工艺
    """
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="工程不存在")

    if not project.start_date:
        raise HTTPException(status_code=400, detail="工程未设置开工日期，无法自动排期")

    # 读取关联工艺
    process_links = (
        db.query(ProjectProcess)
        .filter(ProjectProcess.project_id == project_id)
        .all()
    )
    if not process_links:
        raise HTTPException(status_code=400, detail="工程未关联任何施工工艺，无法排期")

    # 获取工艺详情
    process_ids = [pl.process_id for pl in process_links]
    processes = db.query(Process).filter(Process.id.in_(process_ids)).all()
    process_map = {p.id: p for p in processes}

    # 按类别分组：土建 / 电气
    civil_procs = [p for p in processes if p.category == "土建"]
    electric_procs = [p for p in processes if p.category == "电气"]
    other_procs = [p for p in processes if p.category not in ("土建", "电气")]

    try:
        start_dt = datetime.strptime(project.start_date, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="工程开工日期格式错误")

    # 删除该工程已有的自动排期任务
    db.query(ScheduleTask).filter(ScheduleTask.project_id == project_id).delete()
    db.flush()

    schedule_results = []

    def _schedule_group(proc_list, base_start_dt):
        """对一组工艺进行排期，返回这组工艺的结束时间"""
        if not proc_list:
            return base_start_dt

        # 建立依赖关系图
        # 计算每个工艺的最早开始时间
        scheduled = {}  # process_id -> (start_date, end_date)
        remaining = set(p.id for p in proc_list)
        max_iterations = len(proc_list) + 10  # 防死循环
        iteration = 0

        while remaining and iteration < max_iterations:
            iteration += 1
            progress = False
            for pid in list(remaining):
                p = process_map[pid]
                # 检查依赖是否都已排期
                dep_ids = []
                if p.depends_on:
                    dep_str = str(p.depends_on).strip()
                    if dep_str:
                        dep_ids = [int(d.strip()) for d in dep_str.split(",") if d.strip().isdigit()]

                # 只关注同组内的依赖
                group_ids = set(pp.id for pp in proc_list)
                relevant_deps = [d for d in dep_ids if d in group_ids]

                all_deps_done = all(d in scheduled for d in relevant_deps)
                if not all_deps_done:
                    continue

                # 计算开始时间：所有前置工艺完成后
                if relevant_deps:
                    dep_end = max(scheduled[d][1] for d in relevant_deps)
                    proc_start = dep_end + timedelta(days=1)
                else:
                    proc_start = base_start_dt

                duration = p.duration_days if p.duration_days and p.duration_days > 0 else 1
                proc_end = proc_start + timedelta(days=duration - 1)

                scheduled[pid] = (proc_start, proc_end)
                remaining.discard(pid)
                progress = True

            if not progress:
                # 无法继续排期（可能存在循环依赖），强制排剩余的
                for pid in list(remaining):
                    p = process_map[pid]
                    proc_start = base_start_dt
                    duration = p.duration_days if p.duration_days and p.duration_days > 0 else 1
                    proc_end = proc_start + timedelta(days=duration - 1)
                    scheduled[pid] = (proc_start, proc_end)
                    remaining.discard(pid)
                break

        # 对于无依赖的并行工艺，按 sort_order 排序后重新计算
        # 无依赖的工艺可以并行，取工期最长的作为后续工艺的开始时间
        no_dep_procs = [p for p in proc_list if not p.depends_on or not str(p.depends_on).strip()]
        has_dep_procs = [p for p in proc_list if p.depends_on and str(p.depends_on).strip()]

        # 无依赖工艺并行：它们都从 base_start_dt 开始
        parallel_end = base_start_dt
        for p in no_dep_procs:
            duration = p.duration_days if p.duration_days and p.duration_days > 0 else 1
            p_start = base_start_dt
            p_end = p_start + timedelta(days=duration - 1)
            scheduled[p.id] = (p_start, p_end)
            if p_end > parallel_end:
                parallel_end = p_end

        # 有依赖工艺：重新按依赖计算
        remaining = set(p.id for p in has_dep_procs)
        max_iterations = len(has_dep_procs) + 10
        iteration = 0
        while remaining and iteration < max_iterations:
            iteration += 1
            progress = False
            for pid in list(remaining):
                p = process_map[pid]
                dep_str = str(p.depends_on).strip() if p.depends_on else ""
                dep_ids = [int(d.strip()) for d in dep_str.split(",") if d.strip().isdigit()]
                group_ids = set(pp.id for pp in proc_list)
                relevant_deps = [d for d in dep_ids if d in group_ids]

                all_deps_done = all(d in scheduled for d in relevant_deps)
                if not all_deps_done:
                    continue

                if relevant_deps:
                    dep_end = max(scheduled[d][1] for d in relevant_deps)
                    proc_start = dep_end + timedelta(days=1)
                else:
                    proc_start = parallel_end + timedelta(days=1)

                duration = p.duration_days if p.duration_days and p.duration_days > 0 else 1
                proc_end = proc_start + timedelta(days=duration - 1)
                scheduled[pid] = (proc_start, proc_end)
                remaining.discard(pid)
                progress = True

            if not progress:
                for pid in list(remaining):
                    p = process_map[pid]
                    proc_start = parallel_end + timedelta(days=1)
                    duration = p.duration_days if p.duration_days and p.duration_days > 0 else 1
                    proc_end = proc_start + timedelta(days=duration - 1)
                    scheduled[pid] = (proc_start, proc_end)
                    remaining.discard(pid)
                break

        # 返回组内最晚结束时间
        if scheduled:
            group_end = max(end for _, end in scheduled.values())
        else:
            group_end = base_start_dt

        return scheduled, group_end

    # 先排土建类
    civil_scheduled = {}
    civil_end = start_dt
    if civil_procs:
        civil_scheduled, civil_end = _schedule_group(civil_procs, start_dt)

    # 再排其他类（与土建并行）
    other_scheduled = {}
    other_end = start_dt
    if other_procs:
        other_scheduled, other_end = _schedule_group(other_procs, start_dt)

    # 电气类：土建全部完成后开始
    electric_scheduled = {}
    electric_end = civil_end
    if electric_procs:
        electric_start = civil_end + timedelta(days=1)
        electric_scheduled, electric_end = _schedule_group(electric_procs, electric_start)

    # 合并所有排期
    all_scheduled = {}
    all_scheduled.update(civil_scheduled)
    all_scheduled.update(other_scheduled)
    all_scheduled.update(electric_scheduled)

    # 生成 ScheduleTask 记录
    sort_idx = 0
    for p in sorted(processes, key=lambda x: (x.sort_order or 0, x.id)):
        if p.id not in all_scheduled:
            continue
        proc_start, proc_end = all_scheduled[p.id]
        task = ScheduleTask(
            project_id=project_id,
            task_name=p.name,
            process_id=p.id,
            parent_id=None,
            start_date=proc_start.strftime('%Y-%m-%d'),
            end_date=proc_end.strftime('%Y-%m-%d'),
            actual_start=None,
            actual_end=None,
            progress=0,
            responsible=None,
            sort_order=sort_idx,
        )
        db.add(task)
        schedule_results.append({
            "process_id": p.id,
            "process_name": p.name,
            "category": p.category,
            "start_date": proc_start.strftime('%Y-%m-%d'),
            "end_date": proc_end.strftime('%Y-%m-%d'),
            "duration_days": p.duration_days or 1,
            "depends_on": str(p.depends_on) if p.depends_on else None,
        })
        sort_idx += 1

    db.commit()

    return {
        "message": f"自动排期完成，共 {len(schedule_results)} 项工艺",
        "project_id": project_id,
        "start_date": project.start_date,
        "items": schedule_results,
    }


# ---- 工艺和人员查询接口 ----

@router.get("/api/processes", summary="获取工艺列表")
def get_processes(
    project_type: Optional[str] = Query(None, description="按工程部位筛选: 配电房/电缆通道排管/电缆井"),
    category: Optional[str] = Query(None, description="按分类筛选: 土建/电气"),
    sub_category: Optional[str] = Query(None, description="按工序等级筛选: 一般/特殊"),
    db: Session = Depends(get_db),
):
    """获取施工工艺列表"""
    query = db.query(Process)
    if project_type:
        query = query.filter(Process.project_type == project_type)
    if category:
        query = query.filter(Process.category == category)
    if sub_category:
        query = query.filter(Process.sub_category.contains(sub_category))
    processes = query.order_by(Process.code, Process.id).all()
    return {
        "items": [
            {
                "id": p.id,
                "code": p.code or "",
                "name": p.name,
                "project_type": p.project_type,
                "category": p.category,
                "sub_category": p.sub_category,
                "flow_steps": p.flow_steps,
                "duration_days": p.duration_days,
                "standards": p.standards,
                "equipment": p.equipment,
                "hazards": p.hazards,
                "safety_measures": p.safety_measures,
                "description": p.description,
            }
            for p in processes
        ],
        "total": len(processes),
    }


@router.delete("/api/processes/{process_id}", summary="删除施工工艺")
def delete_process(process_id: int, db: Session = Depends(get_db)):
    """删除指定施工工艺"""
    process = db.query(Process).filter(Process.id == process_id).first()
    if not process:
        raise HTTPException(status_code=404, detail="工艺不存在")
    db.delete(process)
    db.commit()
    return {"message": "删除成功"}


@router.delete("/api/processes", summary="删除全部施工工艺")
def delete_all_processes(db: Session = Depends(get_db)):
    """删除全部施工工艺数据"""
    count = db.query(Process).count()
    if count == 0:
        return {"message": "没有数据需要删除", "deleted": 0}
    db.query(Process).delete()
    db.commit()
    return {"message": f"已删除全部 {count} 条工艺数据", "deleted": count}


class ProcessUpdate(BaseModel):
    """施工工艺更新请求体"""
    code: Optional[str] = None
    name: Optional[str] = None
    project_type: Optional[str] = None
    category: Optional[str] = None
    sub_category: Optional[str] = None
    flow_steps: Optional[str] = None
    duration_days: Optional[int] = None
    standards: Optional[str] = None
    equipment: Optional[str] = None
    hazards: Optional[str] = None
    safety_measures: Optional[str] = None
    description: Optional[str] = None


@router.put("/api/processes/{process_id}", summary="修改施工工艺")
def update_process(process_id: int, data: ProcessUpdate, db: Session = Depends(get_db)):
    """修改单条施工工艺"""
    process = db.query(Process).filter(Process.id == process_id).first()
    if not process:
        raise HTTPException(status_code=404, detail="工艺不存在")
    update_data = data.model_dump(exclude_unset=True)

    if "code" in update_data:
        new_code = (update_data["code"] or "").strip() or None
        if new_code and new_code != process.code:
            conflict = db.query(Process).filter(Process.code == new_code, Process.id != process_id).first()
            if conflict:
                raise HTTPException(status_code=400, detail=f"编号'{new_code}'已被其他工艺占用")
        update_data["code"] = new_code

    for key, value in update_data.items():
        setattr(process, key, value)
    db.commit()
    db.refresh(process)
    return {
        "message": "修改成功",
        "id": process.id,
        "code": process.code or "",
        "name": process.name,
    }


@router.get("/api/processes/export", summary="导出施工工艺到Excel")
def export_processes(db: Session = Depends(get_db)):
    """导出全部施工工艺为Excel文件"""
    processes = db.query(Process).order_by(Process.code, Process.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "施工工艺明细"

    # 表头（编号放在第一列方便回填）
    headers = ["编号", "ID", "工艺名称", "工程部位", "工艺分类", "工序等级", "施工工序", "施工天数", "施工标准", "施工机具", "危险源识别", "安全措施", "工艺说明"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 数据行
    for row_idx, p in enumerate(processes, start=2):
        flow_steps_text = ""
        if p.flow_steps:
            try:
                steps = json.loads(p.flow_steps)
                flow_steps_text = " → ".join(steps)
            except (json.JSONDecodeError, TypeError):
                flow_steps_text = p.flow_steps

        row_data = [
            p.code or "",
            p.id,
            p.name or "",
            p.project_type or "",
            p.category or "",
            p.sub_category or "",
            flow_steps_text,
            p.duration_days if p.duration_days is not None else "",
            p.standards or "",
            p.equipment or "",
            p.hazards or "",
            p.safety_measures or "",
            p.description or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    # 调整列宽
    col_widths = [10, 6, 22, 18, 10, 22, 50, 10, 50, 40, 40, 50, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('施工工艺_导出_' + datetime.now().strftime('%Y%m%d') + '.xlsx')}"},
    )


@router.get("/api/project-types", summary="获取工程部位类别列表")
def get_project_types(db: Session = Depends(get_db)):
    """从施工工艺数据中提取所有不重复的工程部位类别"""
    types = db.query(Process.project_type).filter(Process.project_type.isnot(None), Process.project_type != "").distinct().all()
    return {"items": [t[0] for t in types if t[0]]}


@router.get("/api/workers", summary="获取施工人员列表")
def get_workers(
    role: Optional[str] = Query(None, description="按角色筛选"),
    db: Session = Depends(get_db),
):
    """获取施工人员列表"""
    query = db.query(Worker)
    if role:
        query = query.filter((Worker.role == role) | (Worker.role2 == role) | (Worker.role3 == role))
    workers = query.all()
    return {
        "items": [
            {
                "id": w.id,
                "name": w.name,
                "role": w.role,
                "role2": w.role2,
                "role3": w.role3,
                "team": w.team,
                "certification": w.certification,
                "certification2": w.certification2,
                "certification3": w.certification3,
            }
            for w in workers
        ]
    }


@router.delete("/api/workers/{worker_id}", summary="删除施工人员")
def delete_worker(worker_id: int, db: Session = Depends(get_db)):
    """删除指定施工人员"""
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="人员不存在")
    db.delete(worker)
    db.commit()
    return {"message": "删除成功"}


class WorkerUpdateRequest(BaseModel):
    name: str
    role: Optional[str] = None
    role2: Optional[str] = None
    role3: Optional[str] = None
    team: Optional[str] = None
    certification: Optional[str] = None
    certification2: Optional[str] = None
    certification3: Optional[str] = None


@router.put("/api/workers/{worker_id}", summary="更新施工人员")
def update_worker(worker_id: int, body: WorkerUpdateRequest, db: Session = Depends(get_db)):
    """更新指定施工人员信息"""
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="人员不存在")
    worker.name = body.name
    worker.role = body.role
    worker.role2 = body.role2
    worker.role3 = body.role3
    worker.team = body.team
    worker.certification = body.certification
    worker.certification2 = body.certification2
    worker.certification3 = body.certification3
    db.commit()
    db.refresh(worker)
    return {"message": "更新成功", "worker": WorkerResponse.model_validate(worker)}


@router.post("/api/workers/import", summary="批量导入施工人员")
async def import_workers(
    file: UploadFile = File(..., description="Excel(.xlsx)或CSV文件"),
    overwrite: bool = Query(False, description="是否覆盖已存在的人员（按姓名匹配）"),
    db: Session = Depends(get_db),
):
    """上传施工人员表文件，解析后批量导入数据库

    支持的文件格式：Excel(.xlsx) 或 CSV(.csv)

    表格列要求（第一行为表头）：
    - 姓名（必填）
    - 角色/职务（可选）
    - 角色/职务2（可选）
    - 角色/职务3（可选）
    - 所属班组（可选）
    - 资质证书（可选）
    - 资质证书2（可选）
    - 资质证书3（可选）

    overwrite=False 时：已存在姓名的人员将跳过
    overwrite=True 时：已存在姓名的人员将被更新
    """
    filename = file.filename.lower()

    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 格式文件")

    content = await file.read()

    try:
        rows = []
        if filename.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.reader(text.splitlines())
            rows = list(reader)
        else:
            tmp_path = os.path.join(tempfile.gettempdir(), f"upload_{filename}")
            with open(tmp_path, "wb") as f:
                f.write(content)
            wb = openpyxl.load_workbook(tmp_path, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            wb.close()
            os.remove(tmp_path)

        if len(rows) < 2:
            raise HTTPException(status_code=400, detail="文件至少需要包含表头和一行数据")

        header = [str(h).strip() if h else "" for h in rows[0]]
        col_map = {}
        for i, h in enumerate(header):
            if "姓名" in h or "名字" in h:
                col_map["name"] = i
            elif h in ("角色/职务2", "角色2", "职务2"):
                col_map["role2"] = i
            elif h in ("角色/职务3", "角色3", "职务3"):
                col_map["role3"] = i
            elif "角色" in h or "职务" in h or "工种" in h:
                col_map["role"] = i
            elif "班组" in h:
                col_map["team"] = i
            elif h in ("资质证书2", "证书2"):
                col_map["certification2"] = i
            elif h in ("资质证书3", "证书3"):
                col_map["certification3"] = i
            elif "资质" in h or "证书" in h:
                col_map["certification"] = i

        if "name" not in col_map:
            raise HTTPException(status_code=400, detail='表头中未找到"姓名"列，请检查文件格式')

        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or all(not cell for cell in row):
                continue

            name = str(row[col_map["name"]]).strip() if col_map.get("name") is not None and col_map["name"] < len(row) and row[col_map["name"]] else ""
            if not name:
                skipped += 1
                continue

            def _get_str(key):
                if key not in col_map:
                    return None
                idx = col_map[key]
                val = str(row[idx]).strip() if idx < len(row) and row[idx] else None
                return val or None

            role = _get_str("role")
            role2 = _get_str("role2")
            role3 = _get_str("role3")
            team = _get_str("team")
            certification = _get_str("certification")
            certification2 = _get_str("certification2")
            certification3 = _get_str("certification3")

            existing = db.query(Worker).filter(Worker.name == name).first()
            if existing:
                if overwrite:
                    existing.role = role
                    existing.role2 = role2
                    existing.role3 = role3
                    existing.team = team
                    existing.certification = certification
                    existing.certification2 = certification2
                    existing.certification3 = certification3
                    updated += 1
                else:
                    skipped += 1
                    errors.append(f"第{row_idx}行：人员'{name}'已存在，已跳过")
                continue

            worker = Worker(
                name=name,
                role=role,
                role2=role2,
                role3=role3,
                team=team,
                certification=certification,
                certification2=certification2,
                certification3=certification3,
            )
            db.add(worker)
            imported += 1

        db.commit()

        if overwrite:
            message = f"导入完成：新增 {imported} 条，覆盖 {updated} 条，跳过 {skipped} 条"
        else:
            message = f"导入完成：新增 {imported} 条，跳过 {skipped} 条"

        return {
            "message": message,
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")


@router.get("/api/workers/export", summary="导出施工人员到Excel")
def export_workers(db: Session = Depends(get_db)):
    """导出全部施工人员为Excel文件"""
    workers = db.query(Worker).order_by(Worker.team, Worker.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "施工人员明细"

    headers = ["ID", "姓名", "角色/职务", "角色/职务2", "角色/职务3", "所属班组", "资质证书", "资质证书2", "资质证书3"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, w in enumerate(workers, start=2):
        row_data = [
            w.id,
            w.name or "",
            w.role or "",
            w.role2 or "",
            w.role3 or "",
            w.team or "",
            w.certification or "",
            w.certification2 or "",
            w.certification3 or "",
        ]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)

    col_widths = [6, 14, 18, 18, 18, 18, 30, 30, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('施工人员_导出_' + datetime.now().strftime('%Y%m%d') + '.xlsx')}"},
    )


# ---- 文档模板上传/管理接口 ----

DOC_TYPE_MAP = {
    "construction_design": "施工组织设计模板",
    "survey": "项目勘察单模板",
    "tech_briefing": "技术交底模板",
    "safety_briefing": "安全交底模板",
    "gantt_chart": "施工进度横道图模板",
}


@router.post("/api/templates/upload", summary="上传文档模板")
async def upload_template(
    doc_type: str = Query(..., description="文档类型: construction_design/survey/tech_briefing/safety_briefing"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """上传文档模板文件，支持多种格式"""
    if doc_type not in DOC_TYPE_MAP:
        raise HTTPException(status_code=400, detail=f"不支持的文档类型: {doc_type}")

    # 校验文件格式 - 支持多种文档格式
    allowed_extensions = ['.docx', '.doc', '.xlsx', '.xls', '.md', '.pdf']
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail=f"仅支持以下格式的模板文件: {', '.join(allowed_extensions)}")

    # 保存文件，保留原始扩展名
    file_name = f"{DOC_TYPE_MAP[doc_type]}{file_ext}"
    file_path = os.path.join(TEMPLATES_DIR, file_name)
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # 更新或创建数据库记录
    template = db.query(DocTemplate).filter(DocTemplate.doc_type == doc_type).first()
    if template:
        template.file_name = file.filename
        template.file_path = file_path
        template.uploaded_at = None  # 触发 default 重新赋值
    else:
        template = DocTemplate(doc_type=doc_type, file_name=file.filename, file_path=file_path)
        db.add(template)
    db.commit()

    return {"message": "模板上传成功", "doc_type": doc_type, "file_name": file.filename}


@router.get("/api/templates", summary="获取模板列表")
def get_templates(db: Session = Depends(get_db)):
    """获取所有文档模板信息"""
    templates = db.query(DocTemplate).all()
    result = []
    for doc_type, label in DOC_TYPE_MAP.items():
        tpl = next((t for t in templates if t.doc_type == doc_type), None)
        result.append({
            "doc_type": doc_type,
            "label": label,
            "file_name": tpl.file_name if tpl else None,
            "uploaded_at": str(tpl.uploaded_at) if tpl and tpl.uploaded_at else None,
        })
    return {"items": result}


@router.delete("/api/templates/{doc_type}", summary="删除模板")
def delete_template(doc_type: str, db: Session = Depends(get_db)):
    """删除指定类型的模板"""
    if doc_type not in DOC_TYPE_MAP:
        raise HTTPException(status_code=400, detail=f"不支持的文档类型: {doc_type}")

    template = db.query(DocTemplate).filter(DocTemplate.doc_type == doc_type).first()
    if template:
        # 删除文件
        if template.file_path and os.path.exists(template.file_path):
            os.remove(template.file_path)
        db.delete(template)
        db.commit()

    return {"message": "模板删除成功", "doc_type": doc_type}


# ---- 施工工艺批量导入接口 ----

# 多级表头 Excel 的列布局常量
# 列[0] 序号 | 列[1] 工程部位 | 列[2] 工序名称
# 列[3]~[7]  土建一般: 施工标准/施工工序/设备/危险源/安全措施
# 列[8]      土建特殊工序名
# 列[9]~[13] 土建特殊: 施工标准/施工工序/设备/危险源/安全措施
# 列[14]     电气一般工序名
# 列[15]~[19] 电气一般: 施工标准/施工工序/设备/危险源/安全措施
# 列[20]     电气特殊工序名
# 列[21]~[25] 电气特殊: 施工标准/施工工序/设备/危险源/安全措施

# 每组5列的偏移: [0]施工标准 [1]施工工序 [2]设备 [3]危险源 [4]安全措施
_COL_OFFSETS = {"standards": 0, "flow_steps": 1, "equipment": 2, "hazards": 3, "safety_measures": 4}

# 4个分组的起始列和元信息
_GROUPS = [
    {"start_col": 3, "name_col": 2, "category": "土建", "sub_category": "一般"},
    {"start_col": 9, "name_col": 8, "category": "土建", "sub_category": "特殊（需专项施工方案）"},
    {"start_col": 15, "name_col": 14, "category": "电气", "sub_category": "一般"},
    {"start_col": 21, "name_col": 20, "category": "电气", "sub_category": "特殊（需专项施工方案）"},
]


def _safe_str(val):
    """将单元格值转为字符串，None 返回空串"""
    if val is None:
        return ""
    return str(val).strip()


def _parse_int_cell(val):
    """将单元格值转为整数，无法解析返回 None"""
    if val is None:
        return None
    try:
        # 处理浮点数（Excel 数字列默认为 float）
        f = float(val)
        if f == int(f):
            return int(f)
        return int(f)
    except (ValueError, TypeError):
        return None


def _parse_flow_steps(text):
    """将施工工序文本解析为 JSON 数组字符串，保留→分隔符"""
    if not text:
        return None
    # 按 → 分割为步骤列表
    steps = [s.strip() for s in text.split("→") if s.strip()]
    if not steps:
        return None
    return json.dumps(steps, ensure_ascii=False)


def _category_num(category):
    """工艺分类编号: 土建=1, 电气=2, 其他=0"""
    if not category:
        return 0
    return {"土建": 1, "电气": 2}.get(category, 0)


def _sub_category_num(sub_category):
    """工序等级编号: 一般=0, 特殊=2, 其他=0"""
    if not sub_category:
        return 0
    if "特殊" in sub_category:
        return 2
    return 0


def _format_process_code(category, sub_category, letter):
    """格式化工艺编号: {cat_num}-{sub_num}-{letter}"""
    return f"{_category_num(category)}-{_sub_category_num(sub_category)}-{letter}"


def _generate_process_letter(db, category, sub_category, used_codes=None, exclude_id=None):
    """生成下一个可用的字母 (A-Z, AA-AZ, ...)，used_codes 为已在导入文件中出现的编号集合"""
    prefix = f"{_category_num(category)}-{_sub_category_num(sub_category)}-"
    query = db.query(Process.code).filter(Process.code.like(f"{prefix}%"))
    if exclude_id is not None:
        query = query.filter(Process.id != exclude_id)
    existing_codes = {r[0] for r in query.all() if r[0]}

    if used_codes:
        existing_codes |= used_codes

    used_letters = set()
    for code in existing_codes:
        if code.startswith(prefix):
            tail = code[len(prefix):]
            if tail and tail.isalpha():
                used_letters.add(tail.upper())

    for i in range(26):
        letter = chr(ord('A') + i)
        if letter not in used_letters:
            return letter

    # 超过 26 个，使用 AA, AB, ...
    extra_idx = 0
    while True:
        first = chr(ord('A') + (extra_idx // 26))
        second = chr(ord('A') + (extra_idx % 26))
        letter = f"{first}{second}"
        if letter not in used_letters:
            return letter
        extra_idx += 1


def _generate_process_code(db, category, sub_category, used_codes=None, exclude_id=None):
    """生成下一个可用的工艺完整编号"""
    letter = _generate_process_letter(db, category, sub_category, used_codes, exclude_id)
    return _format_process_code(category, sub_category, letter)


@router.post("/api/processes/import/preview", summary="预览工艺导入")
async def preview_import_processes(
    file: UploadFile = File(..., description="Excel(.xlsx)或CSV文件"),
    db: Session = Depends(get_db),
):
    """解析上传的工艺表文件，返回新增/覆盖候选清单（不写入数据库）

    返回 preview_id，前端确认覆盖时需回传。
    """
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 格式文件")

    content = await file.read()

    try:
        items = _parse_process_file(content, file.filename)

        # 为没有编号的项自动生成编号，并按 (category, sub_category) 组追踪已用编号
        used_codes_by_group: dict = {}
        new_items = []
        overwrite_items = []

        for item in items:
            code = (item.get("code") or "").strip()
            if not code:
                key = (item.get("category") or "", item.get("sub_category") or "")
                used = used_codes_by_group.get(key, set())
                code = _generate_process_code(db, item.get("category"), item.get("sub_category"), used)
                used_codes_by_group[key] = used | {code}
            item["code"] = code

            existing = db.query(Process).filter(Process.code == code).first()
            if existing:
                overwrite_items.append({
                    "row_idx": item.get("row_idx"),
                    "code": code,
                    "name": item.get("name"),
                    "project_type": item.get("project_type"),
                    "category": item.get("category"),
                    "sub_category": item.get("sub_category"),
                    "existing_id": existing.id,
                    "existing_name": existing.name,
                })
            else:
                new_items.append({
                    "row_idx": item.get("row_idx"),
                    "code": code,
                    "name": item.get("name"),
                    "project_type": item.get("project_type"),
                    "category": item.get("category"),
                    "sub_category": item.get("sub_category"),
                })

        preview_id = str(uuid.uuid4())
        _IMPORT_CACHE[preview_id] = {
            "items": items,
            "created_at": datetime.now(),
        }

        return {
            "preview_id": preview_id,
            "new_items": new_items,
            "overwrite_items": overwrite_items,
            "new_count": len(new_items),
            "overwrite_count": len(overwrite_items),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件解析失败: {str(e)}")


class ImportConfirmRequest(BaseModel):
    """工艺导入确认请求体"""
    preview_id: str
    overwrite_codes: List[str] = []


@router.post("/api/processes/import/confirm", summary="确认工艺导入")
async def confirm_import_processes(
    request: ImportConfirmRequest,
    db: Session = Depends(get_db),
):
    """根据预览结果执行导入：overwrite_codes 中的编号将被覆盖更新，其余已存在编号自动跳过"""
    if request.preview_id not in _IMPORT_CACHE:
        raise HTTPException(status_code=400, detail="预览已过期或不存在，请重新上传文件")

    preview_data = _IMPORT_CACHE[request.preview_id]
    items = preview_data["items"]
    overwrite_set = set(request.overwrite_codes or [])

    imported = 0
    updated = 0
    skipped = 0
    errors = []

    for item in items:
        code = item.get("code", "")
        if not code:
            skipped += 1
            continue

        existing = db.query(Process).filter(Process.code == code).first()
        if existing:
            if code in overwrite_set:
                for key in ("name", "project_type", "category", "sub_category",
                           "flow_steps", "duration_days", "standards", "equipment", "hazards", "safety_measures", "description"):
                    setattr(existing, key, item.get(key))
                updated += 1
            else:
                skipped += 1
                errors.append(f"编号'{code}'（{item.get('name')}）已存在，已跳过")
            continue

        process = Process(
            code=code,
            name=item.get("name") or "",
            project_type=item.get("project_type"),
            category=item.get("category"),
            sub_category=item.get("sub_category"),
            flow_steps=item.get("flow_steps"),
            duration_days=item.get("duration_days"),
            standards=item.get("standards"),
            equipment=item.get("equipment"),
            hazards=item.get("hazards"),
            safety_measures=item.get("safety_measures"),
            description=item.get("description"),
        )
        db.add(process)
        imported += 1

    db.commit()
    del _IMPORT_CACHE[request.preview_id]

    return {
        "message": f"导入完成：新增 {imported} 条，覆盖 {updated} 条，跳过 {skipped} 条",
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def _parse_process_file(content, filename):
    """解析上传的工艺表文件，返回工艺项目列表（不含 code 字段，将由 preview 端补全）"""
    filename_lower = filename.lower()
    rows = []
    if filename_lower.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.reader(text.splitlines())
        rows = list(reader)
    else:
        tmp_path = os.path.join(tempfile.gettempdir(), f"upload_{filename_lower}")
        with open(tmp_path, "wb") as f:
            f.write(content)
        wb = openpyxl.load_workbook(tmp_path, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        os.remove(tmp_path)

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="文件至少需要包含表头和一行数据")

    row1_values = [str(c).strip() for c in rows[0] if c]
    row1_text = " ".join(row1_values)
    row3_text = " ".join(str(c) for c in rows[2] if c) if len(rows) > 2 else ""

    # 多级表头检测：
    # 1) 第1行包含"土建"或"电气"作为分组标题（独立值）
    # 2) 第3行包含"工艺名称"作为列标题
    has_group_headers = "土建" in row1_values or "电气" in row1_values
    has_column_header_in_row3 = "工艺名称" in row3_text
    if has_group_headers or has_column_header_in_row3:
        return _parse_multi_header(rows)
    return _parse_simple_header(rows)


def _parse_multi_header(rows):
    """解析多级表头 Excel（施工工艺_完善审计版格式），返回工艺项目列表"""
    items = []
    current_project_type = ""

    for row_idx, row in enumerate(rows[3:], start=4):
        if not row or all(_safe_str(c) == "" for c in row):
            continue

        col1 = _safe_str(row[1]) if len(row) > 1 else ""
        if col1:
            current_project_type = col1
        if not current_project_type:
            continue

        for group in _GROUPS:
            name_col = group["name_col"]
            start_col = group["start_col"]
            name = _safe_str(row[name_col]) if len(row) > name_col else ""
            if not name:
                continue

            standards = _safe_str(row[start_col + 0]) if len(row) > start_col + 0 else ""
            flow_steps_raw = _safe_str(row[start_col + 1]) if len(row) > start_col + 1 else ""
            equipment = _safe_str(row[start_col + 2]) if len(row) > start_col + 2 else ""
            hazards = _safe_str(row[start_col + 3]) if len(row) > start_col + 3 else ""
            safety_measures = _safe_str(row[start_col + 4]) if len(row) > start_col + 4 else ""

            items.append({
                "row_idx": row_idx,
                "code": "",
                "name": name,
                "project_type": current_project_type,
                "category": group["category"],
                "sub_category": group["sub_category"],
                "flow_steps": _parse_flow_steps(flow_steps_raw),
                "standards": standards or None,
                "equipment": equipment or None,
                "hazards": hazards or None,
                "safety_measures": safety_measures or None,
                "description": None,
            })
    return items


def _parse_simple_header(rows):
    """解析标准单行表头格式，返回工艺项目列表"""
    header = [str(h).strip() if h else "" for h in rows[0]]
    col_map = {}
    for i, h in enumerate(header):
        if "编号" in h or "code" in h.lower():
            col_map["code"] = i
        elif "工艺名称" in h or "名称" in h:
            col_map["name"] = i
        elif "工程部位" in h or "部位" in h:
            col_map["project_type"] = i
        elif "分类" in h:
            col_map["category"] = i
        elif "等级" in h or "子分类" in h:
            col_map["sub_category"] = i
        elif "流程" in h or "工序" in h:
            col_map["flow_steps"] = i
        elif "施工天数" in h or "工期" in h or "天数" in h:
            col_map["duration_days"] = i
        elif "标准" in h:
            col_map["standards"] = i
        elif "设备" in h or "机具" in h:
            col_map["equipment"] = i
        elif "危险源" in h or "危险" in h:
            col_map["hazards"] = i
        elif "安全措施" in h or "安全" in h:
            col_map["safety_measures"] = i
        elif "说明" in h or "描述" in h:
            col_map["description"] = i

    if "name" not in col_map:
        raise HTTPException(status_code=400, detail='表头中未找到"工艺名称"列，请检查文件格式')

    items = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(not cell for cell in row):
            continue
        name = _safe_str(row[col_map["name"]]) if col_map.get("name") is not None and col_map["name"] < len(row) and row[col_map["name"]] else ""
        if not name:
            continue

        def _get_col(key):
            if key not in col_map:
                return None
            idx = col_map[key]
            val = _safe_str(row[idx]) if idx < len(row) and row[idx] else ""
            return val or None

        items.append({
            "row_idx": row_idx,
            "code": _get_col("code") or "",
            "name": name,
            "project_type": _get_col("project_type"),
            "category": _get_col("category"),
            "sub_category": _get_col("sub_category"),
            "flow_steps": _parse_flow_steps(_get_col("flow_steps") or ""),
            "duration_days": _parse_int_cell(row[col_map["duration_days"]]) if "duration_days" in col_map and col_map["duration_days"] < len(row) else None,
            "standards": _get_col("standards"),
            "equipment": _get_col("equipment"),
            "hazards": _get_col("hazards"),
            "safety_measures": _get_col("safety_measures"),
            "description": _get_col("description"),
        })
    return items


@router.post("/api/processes/import", summary="批量导入施工工艺（兼容旧版）", deprecated=True)
async def import_processes(
    file: UploadFile = File(..., description="Excel(.xlsx)或CSV文件"),
    overwrite: bool = Query(False, description="是否覆盖已存在的工艺（默认跳过）"),
    db: Session = Depends(get_db),
):
    """兼容旧版的导入接口：内部走 preview + confirm 流程
    overwrite=false 时：新编号导入，同编号跳过
    overwrite=true 时：新编号导入，同编号覆盖
    """
    filename = file.filename.lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xls") or filename.endswith(".csv")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls / .csv 格式文件")

    content = await file.read()

    try:
        items = _parse_process_file(content, file.filename)

        used_codes_by_group: dict = {}
        preview_items = []
        for item in items:
            code = (item.get("code") or "").strip()
            if not code:
                key = (item.get("category") or "", item.get("sub_category") or "")
                used = used_codes_by_group.get(key, set())
                code = _generate_process_code(db, item.get("category"), item.get("sub_category"), used)
                used_codes_by_group[key] = used | {code}
            item["code"] = code
            preview_items.append(item)

        preview_id = str(uuid.uuid4())
        _IMPORT_CACHE[preview_id] = {
            "items": preview_items,
            "created_at": datetime.now(),
        }

        overwrite_codes = []
        if overwrite:
            for item in preview_items:
                code = item.get("code", "")
                if db.query(Process).filter(Process.code == code).first():
                    overwrite_codes.append(code)

        req = ImportConfirmRequest(preview_id=preview_id, overwrite_codes=overwrite_codes)
        return await confirm_import_processes(req, db)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败：{str(e)}")
